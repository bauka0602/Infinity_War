import base64
import hashlib
import re
from io import BytesIO

from sqlalchemy import case, delete, func, or_, select, update

from ..auth.service import require_auth_user
from ..core.config import DB_LOCK, TEACHER_EMAIL_DOMAIN
from ..core.errors import ApiError
from ..core.orm import SessionLocal
from ..models import Course, CourseComponent, Group, IupEntry, Schedule, Section, Teacher
from ..sections.lesson_rules import requires_computers_for_component
from ..teachers.utils import build_teacher_name_signature, normalize_teacher_name

COURSE_EDUCATIONAL_PROGRAMME_GROUP_ALIASES = {
    "b057": "B057 - Информационные технологии",
    "b057 информационные технологии": "B057 - Информационные технологии",
    "b057 - информационные технологии": "B057 - Информационные технологии",
    "информационные технологии": "B057 - Информационные технологии",
}

IUP_COMPONENT_CODES = {"ОК", "ВК", "КВ", "ДВО", "УПП"}
IUP_LESSON_PHRASES = {
    "Лекции": "lecture",
    "Практики, Семинары": "practical",
    "Лабораторные работы": "lab",
    "Самостоятельная работа студента и преподавателя": "srop",
    "Учебная практика": "practice",
    "Производственная практика": "practice",
}
IUP_ACTIVE_LESSON_TYPES = {"lecture", "practical", "lab"}

ROP_PERIOD_COLUMN_GROUPS = (
    {
        "academic_period_column": 10,
        "total": 10,
        "lecture": 11,
        "practical": 12,
        "lab": 13,
        "studio": 14,
        "practice": 15,
        "srop": 16,
        "sro": 17,
    },
    {
        "academic_period_column": 18,
        "total": 18,
        "lecture": 19,
        "practical": 20,
        "lab": 21,
        "studio": 22,
        "practice": 23,
        "srop": 24,
        "sro": 25,
    },
)

ROP_LESSON_TYPES = ("lecture", "practical", "lab", "studio", "practice", "srop")


def _load_workbook(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel import dependency is not installed on the server.",
        ) from exc

    try:
        return load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ApiError(
            400,
            "bad_request",
            "Не удалось прочитать Excel файл. Используйте формат .xlsx.",
        ) from exc


def _decode_file_payload(payload, allowed_extensions, format_message):
    file_name = (payload.get("fileName") or "").strip()
    file_content = payload.get("fileContent")

    if not file_name or not file_content:
        raise ApiError(
            400,
            "fill_required_fields",
            "Заполните поля: fileName, fileContent",
            {"fields": ["fileName", "fileContent"]},
        )

    if not file_name.lower().endswith(tuple(allowed_extensions)):
        raise ApiError(
            400,
            "bad_request",
            format_message,
        )

    if "," in file_content:
        file_content = file_content.split(",", 1)[1]

    try:
        return file_name, base64.b64decode(file_content)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Некорректное содержимое файла.") from exc


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ").replace("-", "_")


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def _cell_text(value):
    value = _normalize_cell(value)
    return str(value).strip() if value not in (None, "") else ""


def _number_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    normalized = str(value).replace(",", ".").strip()
    if not normalized:
        return None
    try:
        parsed = float(normalized)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def _safe_row_value(row, index):
    return row[index] if index < len(row) else ""


def _normalize_course_department(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return COURSE_EDUCATIONAL_PROGRAMME_GROUP_ALIASES.get(_normalize_header(text), text)

def _upsert_rop_course(connection, course, offering):
    programme = course.get("programme") or ""
    description = (
        f"Imported from ROP. Component: {course.get('component') or '-'}; "
        f"cycle: {course.get('cycle') or '-'}; academic period: {offering['academicPeriod']}."
    )
    with SessionLocal() as session:
        existing = session.scalar(
            select(Course).where(
                func.lower(Course.code) == func.lower(course["code"]),
                func.lower(Course.name) == func.lower(course["name"]),
                Course.semester == offering["academicPeriod"],
                Course.year == course.get("studyYear"),
                func.lower(Course.programme) == func.lower(programme),
            )
        )
        result = "updated" if existing else "inserted"
        row = existing or Course(instructor_id=None, instructor_name="")
        row.name = course["name"]
        row.code = course["code"]
        row.credits = course.get("credits")
        row.hours = offering.get("totalHours")
        row.description = description
        row.year = course.get("studyYear")
        row.semester = offering["academicPeriod"]
        row.department = _normalize_course_department(course.get("department") or "B057")
        row.programme = programme
        row.module_type = course.get("moduleType") or ""
        row.module_name = course.get("moduleName") or ""
        row.cycle = course.get("cycle") or ""
        row.component = course.get("component") or ""
        row.language = course.get("language") or ""
        row.academic_year = course.get("academicYear") or ""
        row.entry_year = course.get("entryYear") or ""
        row.requires_computers = 0
        if existing is None:
            session.add(row)
        session.commit()
        return result, row.id


def _component_has_teacher(component):
    return bool(component and component.get("teacher_id"))


def _component_teacher(component):
    if not _component_has_teacher(component):
        return None, ""
    return component.get("teacher_id"), component.get("teacher_name", "")


def _find_rop_component_teacher(connection, course_id, component):
    with SessionLocal() as session:
        exact = session.scalar(
            select(CourseComponent)
            .where(
                CourseComponent.course_id == course_id,
                CourseComponent.lesson_type == component["lessonType"],
                CourseComponent.academic_period == component["academicPeriod"],
                CourseComponent.teacher_id.is_not(None),
            )
            .order_by(CourseComponent.id)
            .limit(1)
        )
        same_lesson_type = session.scalar(
            select(CourseComponent)
            .where(
                CourseComponent.course_id == course_id,
                CourseComponent.lesson_type == component["lessonType"],
                CourseComponent.teacher_id.is_not(None),
            )
            .order_by(CourseComponent.academic_period.is_(None), CourseComponent.id)
            .limit(1)
        )
    if exact:
        return exact.teacher_id, exact.teacher_name or ""
    return (
        (same_lesson_type.teacher_id, same_lesson_type.teacher_name or "")
        if same_lesson_type
        else (None, "")
    )


def _sync_rop_course_components(connection, course_id, course, offering, lesson_components):
    seen_components = set()

    inserted = 0
    updated = 0
    for component in lesson_components:
        if (
            component["courseCode"] != course["code"]
            or component["courseName"] != course["name"]
            or component["academicPeriod"] != offering["academicPeriod"]
        ):
            continue

        component_key = (component["lessonType"], component["academicPeriod"])
        if component_key in seen_components:
            continue
        seen_components.add(component_key)

        with SessionLocal() as session:
            existing_component = session.scalar(
                select(CourseComponent)
                .where(
                    CourseComponent.course_id == course_id,
                    CourseComponent.lesson_type == component["lessonType"],
                    CourseComponent.academic_period == component["academicPeriod"],
                )
                .order_by(CourseComponent.id)
                .limit(1)
            )

        teacher_id, teacher_name = (
            (existing_component.teacher_id, existing_component.teacher_name or "")
            if existing_component and existing_component.teacher_id
            else (None, "")
        )
        if not teacher_id:
            teacher_id, teacher_name = _find_rop_component_teacher(connection, course_id, component)

        with SessionLocal() as session:
            row = session.get(CourseComponent, existing_component.id) if existing_component else None
            if row:
                row.course_code = component["courseCode"]
                row.course_name = component["courseName"]
                row.programme = course.get("programme") or ""
                row.study_year = course.get("studyYear")
                row.semester = component["semester"]
                row.hours = component["hours"]
                row.weekly_classes = component["weeklyClasses"]
                row.requires_computers = 1 if component.get("requiresComputers") else 0
                if row.teacher_id is None:
                    row.teacher_id = teacher_id
                    if not row.teacher_name:
                        row.teacher_name = teacher_name
                session.commit()
                updated += 1
                continue

            row = CourseComponent(
                course_id=course_id,
                course_code=component["courseCode"],
                course_name=component["courseName"],
                programme=course.get("programme") or "",
                study_year=course.get("studyYear"),
                academic_period=component["academicPeriod"],
                semester=component["semester"],
                lesson_type=component["lessonType"],
                hours=component["hours"],
                weekly_classes=component["weeklyClasses"],
                requires_computers=1 if component.get("requiresComputers") else 0,
                teacher_id=teacher_id,
                teacher_name=teacher_name,
            )
            session.add(row)
            session.commit()
            inserted += 1

    return {"inserted": inserted, "updated": updated}


def _load_rop_rows(file_name, file_bytes):
    lower_name = file_name.lower()
    if lower_name.endswith(".xls"):
        try:
            import xlrd
        except ImportError as exc:
            raise ApiError(
                500,
                "internal_server_error",
                "Excel .xls import dependency is not installed on the server.",
            ) from exc

        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
        except Exception as exc:
            raise ApiError(400, "bad_request", "Не удалось прочитать РОП Excel файл.") from exc
        if not workbook.nsheets:
            raise ApiError(400, "bad_request", "В РОП Excel файле нет листов.")
        sheet = workbook.sheet_by_index(0)
        return [
            [_normalize_cell(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel import dependency is not installed on the server.",
        ) from exc

    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Не удалось прочитать РОП Excel файл.") from exc
    sheet = workbook.worksheets[0]
    return [[_normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]


def _extract_quoted_text(text):
    match = re.search(r"[“\"]([^“”\"]+)[”\"]", text)
    return match.group(1).strip() if match else ""


def _extract_academic_year(rows):
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        match = re.search(r"(20\d{2})\s*[-–]\s*(20\d{2})", text)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
    return ""


def _extract_entry_year(rows):
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        if "Год поступления" in text or "Оқуға түскен жылы" in text:
            return text.split(":", 1)[-1].strip()
    return ""


def _extract_programme_name(rows):
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        quoted = _extract_quoted_text(text)
        if quoted and (
            "образовательной программы" in text.lower()
            or "білім беру бағдарламасының" in text.lower()
        ):
            return quoted
    return ""


def _programme_from_file_name(file_name):
    normalized = re.sub(r"[\s_\-]+", " ", str(file_name or "").lower())
    if "сопр" in normalized:
        return "Компьютерная инженерия (СОПР)"
    if re.search(r"(^|[^a-zа-я])(?:би|bi)([^a-zа-я]|$)", normalized):
        return "Бизнес-информатика"
    if re.search(r"(^|[^a-zа-я])(?:ки|ki)([^a-zа-я]|$)", normalized):
        return "Компьютерная инженерия"
    return ""


def _normalize_rop_programme_name(file_name, programme):
    file_programme = _programme_from_file_name(file_name)
    if file_programme:
        return file_programme
    if "сопр" in file_name.lower() and programme:
        return f"{programme} (СОПР)"
    return programme


def _extract_language(file_name, rows):
    lower_name = file_name.lower()
    if "каз" in lower_name or "_kk" in lower_name or "қаз" in lower_name:
        return "kk"
    if "рус" in lower_name or "_ru" in lower_name:
        return "ru"
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        if any(marker in text for marker in ("ЖҰМЫС ОҚУ ЖОСПАРЫ", "Оқуға түскен жылы")):
            return "kk"
    return "ru"


def _extract_study_year(file_name, academic_periods):
    match = re.search(r"[_\-\s](\d)(?:[_\-\s]|$)", file_name)
    if match:
        return int(match.group(1))
    for academic_period in academic_periods:
        period_number = _extract_period_number(academic_period)
        if period_number:
            return int((period_number + 1) / 2)
    return None


def _extract_period_number(value):
    match = re.search(r"\d+", _cell_text(value))
    return int(match.group(0)) if match else None


def _semester_in_year(academic_period):
    if not academic_period:
        return None
    return 1 if academic_period % 2 == 1 else 2


def _weekly_classes_from_hours(hours):
    numeric_hours = _number_or_none(hours)
    if not numeric_hours:
        return 0
    return max(1, int(round(float(numeric_hours) / 15)))


def _find_rop_table_header_row(rows):
    for index, row in enumerate(rows):
        normalized_values = {_normalize_header(value) for value in row}
        if (
            "код дисциплины" in normalized_values
            or "пәннің коды" in normalized_values
            or "пәннің_коды" in normalized_values
        ):
            return index
    raise ApiError(400, "bad_request", "В РОП не найдена таблица дисциплин.")


def _is_rop_course_row(row):
    first_cell = _cell_text(_safe_row_value(row, 0)).lower()
    code = _cell_text(_safe_row_value(row, 5))
    name = _cell_text(_safe_row_value(row, 6))
    if not code or not name:
        return False
    if first_cell.startswith(("итого", "средняя", "барлығы", "оқу жоспары", "орташа")):
        return False
    return True


def _decode_iup_payload(payload):
    return _decode_file_payload(
        payload,
        (".pdf", ".xls", ".xlsx"),
        "Поддерживаются PDF или Excel файлы ИУП формата .pdf, .xls, .xlsx.",
    )


def _extract_pdf_text(file_bytes):
    try:
        import fitz
    except ImportError as exc:
        raise ApiError(
            500,
            "iup_pdf_dependency_missing",
            "Для импорта PDF ИУП нужно установить pymupdf.",
        ) from exc

    try:
        document = fitz.open(stream=file_bytes, filetype="pdf")
        return "\n".join(page.get_text("text") for page in document)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Не удалось прочитать PDF ИУП.") from exc


def _extract_iup_excel_text(file_name, file_bytes):
    rows = _load_rop_rows(file_name, file_bytes)
    lines = []
    for row in rows:
        for value in row:
            text = _cell_text(value)
            if text:
                lines.extend(part.strip() for part in text.splitlines() if part.strip())
    return "\n".join(lines)


def _iup_clean_lines(text):
    return [
        line.replace("\xa0", " ").strip()
        for line in text.splitlines()
        if line.replace("\xa0", " ").strip()
    ]


def _join_wrapped_text(lines):
    result = " ".join(lines)
    result = re.sub(r"-\s+", "", result)
    return re.sub(r"\s+", " ", result).strip()


def _normalise_iup_programme(value):
    text = str(value or "").strip()
    normalized_text = text.lower()
    if "бизнес" in normalized_text or "6b06102" in normalized_text:
        return "Бизнес-информатика"
    is_computer_engineering = "компьютер" in normalized_text or "6b06103" in normalized_text
    if "сопр" in normalized_text or (
        is_computer_engineering
        and re.search(r"\b3\s*(?:г|год|года|лет)\b", normalized_text)
    ):
        return "Компьютерная инженерия (СОПР)"
    if is_computer_engineering:
        return "Компьютерная инженерия"
    return text


def _normalise_iup_language(value):
    text = str(value or "").strip().lower()
    if text in {"kk", "kz", "қазақ", "казахский"}:
        return "kk"
    if text in {"ru", "русский", "орыс"}:
        return "ru"
    if text.startswith(("каз", "қаз")):
        return "kk"
    if text.startswith(("рус", "орыс")):
        return "ru"
    return "ru"


def _infer_iup_group_name(file_name, programme):
    full_match = re.search(r"05[-_]?057[-_](\d{2})[-_](\d{2})", file_name)
    if full_match:
        base_name = f"05-057-{full_match.group(1)}-{full_match.group(2)}"
        if "сопр" in str(programme or "").lower() or "сопр" in file_name.lower():
            return f"{base_name} СОПР"
        return base_name
    match = re.search(r"(\d{2})[-_](\d{2})", file_name)
    if not match:
        return ""
    base_name = f"05-057-{match.group(1)}-{match.group(2)}"
    if "сопр" in str(programme or "").lower() or "сопр" in file_name.lower():
        return f"{base_name} СОПР"
    return base_name


def _extract_iup_metadata(file_name, lines):
    metadata = {
        "fileName": file_name,
        "groupName": "",
        "programme": "",
        "studyCourse": None,
        "language": "ru",
        "academicYear": "",
    }
    for index, line in enumerate(lines):
        lower_line = line.lower()
        if line.startswith("Курс "):
            match = re.search(r"\d+", line)
            metadata["studyCourse"] = int(match.group(0)) if match else None
        elif line.startswith("Язык обучения"):
            metadata["language"] = _normalise_iup_language(line.replace("Язык обучения", "", 1))
        elif line.startswith("Оқыту тілі"):
            metadata["language"] = _normalise_iup_language(line.replace("Оқыту тілі", "", 1))
        elif (
            not metadata["academicYear"]
            and "учебный год" in line.lower()
            and re.search(r"20\d{2}\s*[-–]\s*20\d{2}", line)
        ):
            match = re.search(r"(20\d{2})\s*[-–]\s*(20\d{2})", line)
            metadata["academicYear"] = f"{match.group(1)}-{match.group(2)}"
        elif "(6B" in line or "(7M" in line:
            if "информатика" in line.lower() or "инженерия" in line.lower():
                metadata["programme"] = _normalise_iup_programme(line)
        elif (
            metadata["programme"] == "Компьютерная инженерия"
            and "форма обучения" in lower_line
            and re.search(r"\b3\s*(?:г|год|года|лет)\b", lower_line)
        ):
            metadata["programme"] = "Компьютерная инженерия (СОПР)"

    if not metadata["programme"]:
        metadata["programme"] = _programme_from_file_name(file_name)
    metadata["groupName"] = _infer_iup_group_name(file_name, metadata["programme"])
    return metadata


def _iup_lesson_at(lines, index):
    for size in range(1, 4):
        phrase = " ".join(lines[index : index + size])
        lesson_type = IUP_LESSON_PHRASES.get(phrase)
        if lesson_type:
            return lesson_type, size, phrase
    return None, 0, ""


def _is_iup_course_start(lines, index):
    return (
        index + 1 < len(lines)
        and re.fullmatch(r"\d{1,3}", lines[index] or "")
        and lines[index + 1] in IUP_COMPONENT_CODES
    )


def _split_iup_code_and_name(values):
    code_lines = []
    cursor = 0
    while cursor < len(values):
        line = values[cursor]
        if re.search(r"\d", line):
            code_lines.append(line)
            cursor += 1
            continue
        if (
            not code_lines
            and re.fullmatch(r"[A-Za-z]{2,12}", line)
            and cursor + 1 < len(values)
            and re.fullmatch(r"\d{4}", values[cursor + 1])
        ):
            code_lines.append(line)
            cursor += 1
            continue
        break
    return _join_wrapped_text(code_lines), _join_wrapped_text(values[cursor:])


def _parse_iup_course_block(block):
    if len(block) < 5:
        return None
    course_index = int(block[0])
    component = block[1]
    credits_index = None
    for index in range(2, len(block) - 1):
        if re.fullmatch(r"\d{1,2}", block[index]) and _iup_lesson_at(block, index + 1)[0]:
            credits_index = index
            break
    if credits_index is None:
        return None

    course_code, course_name = _split_iup_code_and_name(block[2:credits_index])
    if not course_code or not course_name:
        return None

    lesson_items = []
    cursor = credits_index + 1
    while cursor < len(block):
        lesson_type, size, phrase = _iup_lesson_at(block, cursor)
        if not lesson_type:
            cursor += 1
            continue
        cursor += size
        teacher_lines = []
        hours = None
        while cursor < len(block):
            hours_match = re.match(r"^(\d+)\b", block[cursor])
            if hours_match:
                hours = int(hours_match.group(1))
                cursor += 1
                break
            if _iup_lesson_at(block, cursor)[0]:
                break
            teacher_lines.append(block[cursor])
            cursor += 1

        teacher_name = _join_wrapped_text(teacher_lines)
        if teacher_name and hours is not None:
            lesson_items.append(
                {
                    "lessonType": lesson_type,
                    "lessonLabel": phrase,
                    "teacherName": teacher_name,
                    "hours": hours,
                }
            )

    return {
        "index": course_index,
        "component": component,
        "code": course_code,
        "name": course_name,
        "credits": int(block[credits_index]),
        "lessonItems": lesson_items,
    }


def _parse_iup_file(file_name, file_bytes):
    lower_name = file_name.lower()
    if lower_name.endswith(".pdf"):
        raw_text = _extract_pdf_text(file_bytes)
    elif lower_name.endswith((".xls", ".xlsx")):
        raw_text = _extract_iup_excel_text(file_name, file_bytes)
    else:
        raise ApiError(400, "bad_request", "Поддерживаются PDF или Excel файлы ИУП.")

    lines = _iup_clean_lines(raw_text)
    metadata = _extract_iup_metadata(file_name, lines)
    courses = []
    entries = []
    current_study_year = metadata.get("studyCourse")
    current_academic_period = None
    index = 0

    while index < len(lines):
        course_year_match = re.match(r"^(\d+)\s+Курс обучения", lines[index])
        if course_year_match:
            current_study_year = int(course_year_match.group(1))

        period_match = re.match(r"^(\d+)\s+Академический период", lines[index])
        if period_match:
            current_academic_period = int(period_match.group(1))

        if not _is_iup_course_start(lines, index):
            index += 1
            continue

        next_index = index + 2
        while next_index < len(lines):
            if _is_iup_course_start(lines, next_index):
                break
            if re.match(r"^(\d+)\s+Курс обучения", lines[next_index]):
                break
            if re.match(r"^(\d+)\s+Академический период", lines[next_index]):
                break
            next_index += 1

        course = _parse_iup_course_block(lines[index:next_index])
        if course:
            course["studyYear"] = current_study_year
            course["academicPeriod"] = current_academic_period
            course["semester"] = current_academic_period
            courses.append(course)
            for lesson in course["lessonItems"]:
                entries.append(
                    {
                        **metadata,
                        "studyYear": current_study_year,
                        "academicPeriod": current_academic_period,
                        "semester": current_academic_period,
                        "component": course["component"],
                        "courseCode": course["code"],
                        "courseName": course["name"],
                        "credits": course["credits"],
                        **lesson,
                    }
                )

        index = next_index

    if not courses:
        raise ApiError(400, "bad_request", "В ИУП не найдены дисциплины.")

    return {
        "metadata": metadata,
        "courses": courses,
        "entries": entries,
        "totals": {
            "courses": len(courses),
            "lessonEntries": len(entries),
            "teachers": len({entry["teacherName"] for entry in entries if entry.get("teacherName")}),
        },
    }


def _teacher_email_from_name(name):
    digest = hashlib.sha1(name.strip().lower().encode("utf-8")).hexdigest()[:12]
    return f"iup-{digest}{TEACHER_EMAIL_DOMAIN}"


def _upsert_iup_teacher(session, teacher_name, language):
    normalized_name = normalize_teacher_name(teacher_name)
    name_signature = build_teacher_name_signature(teacher_name)
    normalized_language = _normalise_iup_language(language)
    existing = session.scalar(
        select(Teacher)
        .where(
            or_(
                func.lower(Teacher.name) == func.lower(teacher_name),
                func.coalesce(Teacher.name_normalized, "") == normalized_name,
                (
                    (func.coalesce(Teacher.name_signature, "") == name_signature)
                    if name_signature
                    else False
                ),
            )
        )
        .order_by(Teacher.id)
        .limit(1)
    )
    if existing:
        current_languages = [
            value.strip().lower()
            for value in str(existing.teaching_languages or "").split(",")
            if value.strip()
        ]
        if normalized_language and normalized_language not in current_languages:
            current_languages.append(normalized_language)
        teaching_languages = ",".join(current_languages or [normalized_language or "ru"])
        if not existing.name_normalized:
            existing.name_normalized = normalized_name
        if not existing.name_signature:
            existing.name_signature = name_signature
        existing.teaching_languages = teaching_languages
        session.flush()
        return existing.id, "existing"

    teacher = Teacher(
        name=teacher_name,
        email=_teacher_email_from_name(teacher_name),
        subject_taught="",
        teaching_languages=normalized_language or "ru",
        name_normalized=normalized_name,
        name_signature=name_signature,
    )
    session.add(teacher)
    session.flush()
    return teacher.id, "inserted"


def _resolve_iup_group_name(session, group_name):
    if not group_name:
        return ""
    exact = session.scalar(select(Group.name).where(Group.name == group_name))
    if exact:
        return exact
    prefixed = session.scalar(
        select(Group.name)
        .where(Group.name.like(f"{group_name}%"))
        .order_by(func.length(Group.name), Group.name)
        .limit(1)
    )
    return prefixed if prefixed else group_name


def _load_group_context(session, group_name):
    if not group_name:
        return None
    row = session.scalar(select(Group).where(Group.name == group_name))
    if row is None:
        return None
    return {
        "name": row.name,
        "programme": row.programme,
        "language": row.language,
        "study_course": row.study_course,
    }


def _find_matching_iup_course(session, course_code, programme, semester):
    programme = programme or ""
    programme_like = f"%{programme}%" if programme else ""
    conditions = [func.lower(Course.code) == func.lower(course_code or "")]
    if programme:
        conditions.append(
            or_(
                func.lower(Course.programme) == func.lower(programme),
                func.lower(Course.programme).like(func.lower(programme_like)),
                Course.programme == "",
            )
        )
    if semester is not None:
        conditions.append(Course.semester == semester)

    return session.scalar(
        select(Course)
        .where(*conditions)
        .order_by(
            case(
                (func.lower(Course.programme) == func.lower(programme), 0),
                (func.lower(Course.programme).like(func.lower(programme_like)), 1),
                (Course.programme == "", 2),
                else_=3,
            ),
            Course.id,
        )
        .limit(1)
    )


def _find_matching_iup_course_relaxed(session, course_code, programme, semester):
    course = _find_matching_iup_course(session, course_code, programme, semester)
    if course:
        return course

    if semester is not None:
        course = session.scalar(
            select(Course)
            .where(
                func.lower(Course.code) == func.lower(course_code or ""),
                Course.semester == semester,
            )
            .order_by(Course.id)
            .limit(1)
        )
        if course:
            return course

    return session.scalar(
        select(Course)
        .where(func.lower(Course.code) == func.lower(course_code or ""))
        .order_by(Course.id)
        .limit(1)
    )


def _course_import_item(code, name, semester=None, programme=None, study_year=None):
    return {
        "code": code or "",
        "name": name or "",
        "semester": semester,
        "programme": programme or "",
        "studyYear": study_year,
    }


def _get_iup_course_lists(session, parsed):
    metadata = parsed["metadata"]
    matched_courses = []
    missing_courses = []
    seen_iup_courses = set()

    for course in parsed["courses"]:
        key = (
            (course.get("code") or "").lower(),
            (course.get("name") or "").lower(),
            course.get("semester"),
        )
        if key in seen_iup_courses:
            continue
        seen_iup_courses.add(key)
        match = _find_matching_iup_course_relaxed(
            session,
            course.get("code", ""),
            metadata.get("programme", ""),
            course.get("semester"),
        )
        item = _course_import_item(
            course.get("code", ""),
            course.get("name", ""),
            course.get("semester"),
            metadata.get("programme", ""),
            course.get("studyYear"),
        )
        if match:
            matched_courses.append(item)
        else:
            missing_courses.append(item)

    return matched_courses, missing_courses


def _missing_courses_for_study_year(missing_courses, study_year):
    return [
        course
        for course in missing_courses
        if int(course.get("studyYear") or 0) == int(study_year)
    ]


def _is_first_year_base_course(course):
    return int(course.get("studyYear") or 0) == 1 and int(course.get("semester") or 0) in {1, 2}


def _missing_first_year_base_courses(missing_courses):
    return [
        course
        for course in missing_courses
        if _is_first_year_base_course(course)
    ]


def _create_iup_missing_courses(session, parsed, missing_courses):
    metadata = parsed["metadata"]
    created = []
    courses_by_key = {
        ((course.get("code") or "").lower(), course.get("semester")): course
        for course in parsed["courses"]
    }

    for item in missing_courses:
        course = courses_by_key.get(((item.get("code") or "").lower(), item.get("semester"))) or item
        course_row = Course(
            name=course.get("name") or item.get("name") or item.get("code") or "Без названия",
            code=course.get("code") or item.get("code") or "",
            credits=course.get("credits"),
            hours=None,
            description=(
                "Imported from IUP: first-year base course not present in ROP."
                if _is_first_year_base_course(course)
                else "Imported from IUP as missing course draft."
            ),
            year=course.get("studyYear"),
            semester=course.get("semester") or item.get("semester"),
            department=_normalize_course_department(metadata.get("educationalProgrammeGroup") or "B057"),
            instructor_id=None,
            instructor_name="",
            programme=metadata.get("programme", ""),
            module_type="",
            module_name="",
            cycle="",
            component=course.get("component", ""),
            language=metadata.get("language", "ru"),
            academic_year=metadata.get("academicYear", ""),
            entry_year="",
            requires_computers=0,
        )
        session.add(course_row)
        session.flush()
        created.append(_course_import_item(
            course.get("code") or item.get("code"),
            course.get("name") or item.get("name"),
            course.get("semester") or item.get("semester"),
            metadata.get("programme", ""),
            course.get("studyYear") or item.get("studyYear"),
        ))

        lesson_entries = [
            entry
            for entry in parsed["entries"]
            if (entry.get("courseCode") or "").lower() == (course.get("code") or item.get("code") or "").lower()
               and entry.get("lessonType") in IUP_ACTIVE_LESSON_TYPES
        ]
        seen_components = set()
        for entry in lesson_entries:
            component_key = (entry.get("lessonType"), entry.get("academicPeriod"))
            if component_key in seen_components:
                continue
            seen_components.add(component_key)
            session.add(
                CourseComponent(
                    course_id=course_row.id,
                    course_code=entry.get("courseCode", ""),
                    course_name=entry.get("courseName", ""),
                    lesson_type=entry.get("lessonType"),
                    hours=entry.get("hours") or 0,
                    weekly_classes=1,
                    academic_period=entry.get("academicPeriod"),
                    semester=entry.get("semester"),
                    requires_computers=1 if requires_computers_for_component(
                        entry.get("lessonType"),
                        entry.get("courseCode", ""),
                        entry.get("courseName", ""),
                        entry.get("studyYear"),
                    ) else 0,
                    teacher_id=None,
                    teacher_name="",
                )
            )

    return created


def _store_iup_entries(connection, parsed, create_missing_courses=False):
    metadata = parsed["metadata"]
    with SessionLocal() as session:
        try:
            metadata["groupName"] = _resolve_iup_group_name(session, metadata.get("groupName", ""))
            group_context = _load_group_context(session, metadata.get("groupName", ""))
            if group_context:
                if group_context.get("language"):
                    metadata["language"] = group_context["language"]
                if group_context.get("study_course") and not metadata.get("studyCourse"):
                    metadata["studyCourse"] = group_context["study_course"]

            session.execute(
                delete(IupEntry).where(
                    IupEntry.file_name == metadata["fileName"],
                    IupEntry.group_name == metadata.get("groupName", ""),
                )
            )

            updated_courses = set()
            updated_components = set()
            teacher_cache = {}
            matched_courses, missing_courses = _get_iup_course_lists(session, parsed)
            created_courses = []
            auto_first_year_courses = _missing_first_year_base_courses(missing_courses)
            courses_to_create = missing_courses if create_missing_courses else auto_first_year_courses
            if courses_to_create:
                created_courses = _create_iup_missing_courses(session, parsed, courses_to_create)
                matched_courses, missing_courses = _get_iup_course_lists(session, parsed)

            for entry in parsed["entries"]:
                teacher_name = entry.get("teacherName", "")
                teacher_id = None
                if teacher_name:
                    if teacher_name not in teacher_cache:
                        teacher_cache[teacher_name] = _upsert_iup_teacher(
                            session,
                            teacher_name,
                            metadata.get("language", "ru"),
                        )
                    teacher_id, _status = teacher_cache[teacher_name]

                session.add(
                    IupEntry(
                        file_name=metadata["fileName"],
                        group_name=metadata.get("groupName", ""),
                        programme=metadata.get("programme", ""),
                        study_course=entry.get("studyYear"),
                        language=metadata.get("language", "ru"),
                        academic_year=metadata.get("academicYear", ""),
                        academic_period=entry.get("academicPeriod"),
                        semester=entry.get("semester"),
                        component=entry.get("component", ""),
                        course_code=entry["courseCode"],
                        course_name=entry["courseName"],
                        credits=entry.get("credits"),
                        lesson_type=entry["lessonType"],
                        teacher_id=teacher_id,
                        teacher_name=teacher_name,
                        hours=entry.get("hours"),
                    )
                )

                if teacher_id and entry["lessonType"] in IUP_ACTIVE_LESSON_TYPES:
                    course = _find_matching_iup_course_relaxed(
                        session,
                        entry["courseCode"],
                        metadata.get("programme", ""),
                        entry.get("semester"),
                    )
                    if course and course.id not in updated_courses:
                        course_update = session.execute(
                            update(Course)
                            .where(
                                Course.id == course.id,
                                or_(
                                    Course.instructor_id.is_(None),
                                    func.coalesce(Course.instructor_name, "") == "",
                                ),
                            )
                            .values(instructor_id=teacher_id, instructor_name=teacher_name)
                        )
                        if max(0, getattr(course_update, "rowcount", 0) or 0):
                            updated_courses.add(course.id)

                    if course:
                        component_conditions = [
                            CourseComponent.course_id == course.id,
                            CourseComponent.lesson_type == entry["lessonType"],
                            or_(
                                CourseComponent.teacher_id.is_(None),
                                func.coalesce(CourseComponent.teacher_name, "") == "",
                            ),
                        ]
                        if entry.get("academicPeriod") is not None:
                            component_conditions.append(
                                CourseComponent.academic_period == entry.get("academicPeriod")
                            )
                        component_update = session.execute(
                            update(CourseComponent)
                            .where(*component_conditions)
                            .values(teacher_id=teacher_id, teacher_name=teacher_name)
                        )
                        updated_count = max(0, getattr(component_update, "rowcount", 0) or 0)
                        if updated_count == 0 and entry.get("academicPeriod") is not None:
                            fallback_update = session.execute(
                                update(CourseComponent)
                                .where(
                                    CourseComponent.course_id == course.id,
                                    CourseComponent.lesson_type == entry["lessonType"],
                                    or_(
                                        CourseComponent.teacher_id.is_(None),
                                        func.coalesce(CourseComponent.teacher_name, "") == "",
                                    ),
                                )
                                .values(teacher_id=teacher_id, teacher_name=teacher_name)
                            )
                            updated_count = max(0, getattr(fallback_update, "rowcount", 0) or 0)
                        session.execute(
                            update(Section)
                            .where(
                                Section.course_id == course.id,
                                Section.lesson_type == entry["lessonType"],
                                or_(
                                    Section.teacher_id.is_(None),
                                    func.coalesce(Section.teacher_name, "") == "",
                                ),
                            )
                            .values(teacher_id=teacher_id, teacher_name=teacher_name)
                        )
                        if updated_count:
                            updated_components.add(
                                (entry["courseCode"], entry["lessonType"], entry.get("academicPeriod"))
                            )

            session.commit()
        except Exception:
            session.rollback()
            raise

    return {
        "iupEntries": len(parsed["entries"]),
        "teachersInserted": sum(1 for _teacher_id, status in teacher_cache.values() if status == "inserted"),
        "teachersExisting": sum(1 for _teacher_id, status in teacher_cache.values() if status == "existing"),
        "coursesUpdated": len(updated_courses),
        "componentsUpdated": len(updated_components),
        "coursesCreated": len(created_courses),
        "coursesExisting": len(matched_courses),
        "coursesMissing": len(missing_courses),
        "courseLists": {
            "inserted": created_courses,
            "existing": matched_courses,
            "missing": missing_courses,
        },
    }


def parse_iup_preview(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    file_name, file_bytes = _decode_iup_payload(payload)
    parsed = _parse_iup_file(file_name, file_bytes)
    with DB_LOCK:
        with SessionLocal() as session:
            matched_courses, missing_courses = _get_iup_course_lists(session, parsed)
    return {
        "metadata": parsed["metadata"],
        "totals": parsed["totals"],
        "courses": [
            {
                "code": course["code"],
                "name": course["name"],
                "component": course["component"],
                "credits": course["credits"],
                "academicPeriod": course.get("academicPeriod"),
            }
            for course in parsed["courses"]
        ],
        "entries": parsed["entries"][:50],
        "courseLists": {
            "existing": matched_courses,
            "missing": missing_courses,
        },
    }


def import_iup_data(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    file_name, file_bytes = _decode_iup_payload(payload)
    parsed = _parse_iup_file(file_name, file_bytes)
    create_missing_courses = bool(payload.get("createMissingCourses"))
    with DB_LOCK:
        stats = _store_iup_entries(None, parsed, create_missing_courses)
    return {
        "success": True,
        "metadata": parsed["metadata"],
        "totals": parsed["totals"],
        "stats": stats,
    }


def parse_rop_preview(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    file_name, file_bytes = _decode_file_payload(
        payload,
        (".xls", ".xlsx"),
        "Поддерживаются Excel файлы РОП формата .xls или .xlsx.",
    )
    rows = _load_rop_rows(file_name, file_bytes)
    header_row_index = _find_rop_table_header_row(rows)
    academic_period_row = rows[header_row_index + 1] if header_row_index + 1 < len(rows) else []
    academic_periods = [
        _extract_period_number(_safe_row_value(academic_period_row, group["academic_period_column"]))
        for group in ROP_PERIOD_COLUMN_GROUPS
    ]

    study_year = _extract_study_year(file_name, academic_periods)
    metadata = {
        "fileName": file_name,
        "programme": _normalize_rop_programme_name(file_name, _extract_programme_name(rows)),
        "academicYear": _extract_academic_year(rows),
        "entryYear": _extract_entry_year(rows),
        "language": _extract_language(file_name, rows),
        "studyYear": study_year,
        "academicPeriods": [period for period in academic_periods if period],
    }

    courses = []
    offerings = []
    lesson_components = []
    seen_courses = set()

    for row_index, row in enumerate(rows[header_row_index + 3 :], start=header_row_index + 4):
        if not _is_rop_course_row(row):
            continue

        course = {
            "rowNumber": row_index,
            "moduleNumber": _cell_text(_safe_row_value(row, 0)),
            "moduleType": _cell_text(_safe_row_value(row, 1)),
            "moduleName": _cell_text(_safe_row_value(row, 2)),
            "cycle": _cell_text(_safe_row_value(row, 3)),
            "component": _cell_text(_safe_row_value(row, 4)),
            "code": _cell_text(_safe_row_value(row, 5)),
            "name": _cell_text(_safe_row_value(row, 6)),
            "credits": _number_or_none(_safe_row_value(row, 7)),
            "examPeriod": _number_or_none(_safe_row_value(row, 9)),
            "programme": metadata["programme"],
            "studyYear": study_year,
            "language": metadata["language"],
            "academicYear": metadata["academicYear"],
            "entryYear": metadata["entryYear"],
        }
        course_key = (course["code"].lower(), course["name"].lower())
        if course_key not in seen_courses:
            seen_courses.add(course_key)
            courses.append(course)

        for group, academic_period in zip(ROP_PERIOD_COLUMN_GROUPS, academic_periods):
            total_hours = _number_or_none(_safe_row_value(row, group["total"]))
            if not academic_period or not total_hours:
                continue

            offering = {
                "courseCode": course["code"],
                "courseName": course["name"],
                "academicPeriod": academic_period,
                "semester": _semester_in_year(academic_period),
                "studyYear": study_year,
                "totalHours": total_hours,
                "lectureHours": _number_or_none(_safe_row_value(row, group["lecture"])) or 0,
                "practicalHours": _number_or_none(_safe_row_value(row, group["practical"])) or 0,
                "labHours": _number_or_none(_safe_row_value(row, group["lab"])) or 0,
                "studioHours": _number_or_none(_safe_row_value(row, group["studio"])) or 0,
                "practiceHours": _number_or_none(_safe_row_value(row, group["practice"])) or 0,
                "sropHours": _number_or_none(_safe_row_value(row, group["srop"])) or 0,
                "sroHours": _number_or_none(_safe_row_value(row, group["sro"])) or 0,
            }
            offerings.append(offering)

            for lesson_type in ROP_LESSON_TYPES:
                hours = offering[f"{lesson_type}Hours"]
                if hours:
                    normalized_lesson_type = "practical" if lesson_type == "studio" else lesson_type
                    lesson_components.append(
                        {
                            "courseCode": course["code"],
                            "courseName": course["name"],
                            "programme": metadata["programme"],
                            "studyYear": study_year,
                            "academicPeriod": academic_period,
                            "semester": offering["semester"],
                            "lessonType": normalized_lesson_type,
                            "hours": hours,
                            "weeklyClasses": _weekly_classes_from_hours(hours),
                            "requiresComputers": requires_computers_for_component(
                                normalized_lesson_type,
                                course["code"],
                                course["name"],
                                study_year,
                            ),
                        }
                    )

    if not courses:
        raise ApiError(400, "bad_request", "В РОП не найдены дисциплины.")

    return {
        "message": "ROP preview parsed successfully.",
        "metadata": metadata,
        "totals": {
            "courses": len(courses),
            "offerings": len(offerings),
            "lessonComponents": len(lesson_components),
        },
        "courses": courses,
        "offerings": offerings,
        "lessonComponents": lesson_components,
    }


def import_rop_data(headers, payload):
    preview = parse_rop_preview(headers, payload)
    course_by_key = {
        (course["code"], course["name"]): course for course in preview["courses"]
    }
    summary = {
        "courses": {"inserted": 0, "updated": 0},
        "courseComponents": {"inserted": 0, "updated": 0},
    }
    course_lists = {
        "inserted": [],
        "existing": [],
    }
    seen_course_results = set()

    with DB_LOCK:
        for offering in preview["offerings"]:
            course = course_by_key.get((offering["courseCode"], offering["courseName"]))
            if not course:
                continue
            result, course_id = _upsert_rop_course(None, course, offering)
            summary["courses"][result] += 1
            list_key = (
                result,
                course["code"].lower(),
                course["name"].lower(),
                offering["academicPeriod"],
                (course.get("programme") or "").lower(),
            )
            if list_key not in seen_course_results:
                seen_course_results.add(list_key)
                course_lists["inserted" if result == "inserted" else "existing"].append(
                    _course_import_item(
                        course["code"],
                        course["name"],
                        offering["academicPeriod"],
                        course.get("programme") or "",
                    )
                )
            component_summary = _sync_rop_course_components(
                None,
                course_id,
                course,
                offering,
                preview["lessonComponents"],
            )
            summary["courseComponents"]["inserted"] += component_summary["inserted"]
            summary["courseComponents"]["updated"] += component_summary["updated"]

    return {
        "message": "ROP import completed successfully.",
        "metadata": preview["metadata"],
        "summary": summary,
        "courseLists": course_lists,
        "totals": {
            "inserted": summary["courses"]["inserted"],
            "updated": summary["courses"]["updated"],
            "courseComponents": summary["courseComponents"]["inserted"],
            "courses": preview["totals"]["courses"],
            "offerings": preview["totals"]["offerings"],
            "lessonComponents": preview["totals"]["lessonComponents"],
        },
    }


EXPORT_TRANSLATIONS = {
    "ru": {
        "schedule": "Расписание",
        "day": "День недели",
        "time": "Время занятия",
        "discipline": "Название предмета",
        "lesson_type": "Тип занятия",
        "teacher": "ФИО преподавателя",
        "room": "Аудитория",
        "monday": "Понедельник",
        "tuesday": "Вторник",
        "wednesday": "Среда",
        "thursday": "Четверг",
        "friday": "Пятница",
        "saturday": "Суббота",
        "sunday": "Воскресенье",
        "lecture": "Лекция",
        "practical": "Практика",
        "lab": "Лаборатория",
    },
    "kk": {
        "schedule": "Кесте",
        "day": "Апта күндері",
        "time": "Сабақ уақыты",
        "discipline": "Пән атауы",
        "lesson_type": "Сабақ түрі",
        "teacher": "Оқытушы аты-жөні",
        "room": "Аудитория",
        "monday": "Дүйсенбі",
        "tuesday": "Сейсенбі",
        "wednesday": "Сәрсенбі",
        "thursday": "Бейсенбі",
        "friday": "Жұма",
        "saturday": "Сенбі",
        "sunday": "Жексенбі",
        "lecture": "Лекция",
        "practical": "Практика",
        "lab": "Зертхана",
    },
    "en": {
        "schedule": "Schedule",
        "day": "Weekday",
        "time": "Lesson time",
        "discipline": "Subject name",
        "lesson_type": "Lesson type",
        "teacher": "Teacher full name",
        "room": "Room",
        "monday": "Monday",
        "tuesday": "Tuesday",
        "wednesday": "Wednesday",
        "thursday": "Thursday",
        "friday": "Friday",
        "saturday": "Saturday",
        "sunday": "Sunday",
        "lecture": "Lecture",
        "practical": "Practical",
        "lab": "Laboratory",
    },
}


def _export_translation(language, key):
    translations = EXPORT_TRANSLATIONS.get(language) or EXPORT_TRANSLATIONS["ru"]
    return translations.get(key) or EXPORT_TRANSLATIONS["ru"].get(key) or key


def _export_weekday_key(day):
    try:
        from datetime import date

        parsed = date.fromisoformat(str(day))
    except (TypeError, ValueError):
        return str(day or "")

    return [
        "monday",
        "tuesday",
        "wednesday",
        "thursday",
        "friday",
        "saturday",
        "sunday",
    ][parsed.weekday()]


def _export_lesson_time(start_hour):
    try:
        hour = int(start_hour)
    except (TypeError, ValueError):
        return ""

    return f"{hour:02d}:00-{hour:02d}:50"


def _export_sheet_title(group_name, used_titles):
    title = re.sub(r"[:\\/?*\[\]]", " ", str(group_name or "Group")).strip()
    title = re.sub(r"\s+", " ", title) or "Group"
    title = title[:31]
    candidate = title
    suffix = 2

    while candidate in used_titles:
        suffix_text = f" {suffix}"
        candidate = f"{title[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_titles.add(candidate)
    return candidate


def generate_schedule_export(headers, semester=None, year=None, language=None, group_id=None):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")
    normalized_language = str(language or "ru").lower()
    if normalized_language not in EXPORT_TRANSLATIONS:
        normalized_language = "ru"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel export dependency is not installed on the server.",
        ) from exc

    with DB_LOCK:
        with SessionLocal() as session:
            statement = (
                select(
                    Schedule.course_name.label("course_name"),
                    Schedule.group_name.label("group_name"),
                    Schedule.subgroup.label("subgroup"),
                    Schedule.teacher_name.label("teacher_name"),
                    Schedule.room_number.label("room_number"),
                    Schedule.day.label("day"),
                    Schedule.start_hour.label("start_hour"),
                    Schedule.semester.label("semester"),
                    Schedule.year.label("year"),
                    Schedule.algorithm.label("algorithm"),
                    Schedule.room_programme.label("room_programme"),
                    Schedule.room_programme_mismatch.label("room_programme_mismatch"),
                    func.coalesce(Section.lesson_type, "lecture").label("lesson_type"),
                )
                .select_from(Schedule)
                .outerjoin(Section, Section.id == Schedule.section_id)
            )
            if semester is not None:
                statement = statement.where(Schedule.semester == semester)
            if year is not None:
                statement = statement.where(Schedule.year == year)
            if group_id is not None:
                statement = statement.where(Schedule.group_id == group_id)
            statement = statement.order_by(
                Schedule.group_name,
                Schedule.day,
                Schedule.start_hour,
                Schedule.course_name,
                Schedule.id,
            )
            schedules = session.execute(statement).mappings().all()

    if not schedules:
        raise ApiError(400, "bad_request", "Расписание ещё не сгенерировано.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    compatibility_sheet = workbook.create_sheet("Schedule")
    compatibility_sheet.append(
        [
            "course_name",
            "group_name",
            "subgroup",
            "teacher_name",
            "room_number",
            "day",
            "start_hour",
            "semester",
            "year",
            "algorithm",
            "room_programme",
            "room_programme_mismatch",
        ]
    )
    for item in schedules:
        compatibility_sheet.append(
            [
                item.get("course_name", ""),
                item.get("group_name", ""),
                item.get("subgroup", ""),
                item.get("teacher_name", ""),
                item.get("room_number", ""),
                item.get("day", ""),
                item.get("start_hour", ""),
                item.get("semester", ""),
                item.get("year", ""),
                item.get("algorithm", ""),
                item.get("room_programme", ""),
                item.get("room_programme_mismatch", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="014531")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="014531", bold=True, size=14)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    text_alignment = Alignment(vertical="top", wrap_text=True)
    used_titles = set()
    schedules_by_group = {}

    for item in schedules:
        schedules_by_group.setdefault(item.get("group_name") or "Group", []).append(item)

    for group_name, group_schedules in schedules_by_group.items():
        sheet = workbook.create_sheet(_export_sheet_title(group_name, used_titles))
        sheet.append([f"{_export_translation(normalized_language, 'schedule')} - {group_name}"])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        sheet["A1"].font = title_font
        sheet["A1"].alignment = center_alignment
        sheet.row_dimensions[1].height = 28
        sheet.append(
            [
                _export_translation(normalized_language, "day"),
                _export_translation(normalized_language, "time"),
                _export_translation(normalized_language, "discipline"),
                _export_translation(normalized_language, "lesson_type"),
                _export_translation(normalized_language, "teacher"),
                _export_translation(normalized_language, "room"),
            ]
        )

        for cell in sheet[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for item in group_schedules:
            weekday_key = _export_weekday_key(item.get("day"))
            day_label = _export_translation(normalized_language, weekday_key)
            time_label = _export_lesson_time(item.get("start_hour"))
            subgroup = str(item.get("subgroup") or "").strip()
            course_name = item.get("course_name") or ""

            if subgroup:
                course_name = f"{course_name} ({subgroup})"

            sheet.append(
                [
                    day_label,
                    time_label,
                    course_name,
                    _export_translation(
                        normalized_language,
                        str(item.get("lesson_type") or "lecture").lower(),
                    ),
                    item.get("teacher_name", ""),
                    item.get("room_number", ""),
                ]
            )

        for column, width in {
            "A": 22,
            "B": 18,
            "C": 46,
            "D": 20,
            "E": 36,
            "F": 18,
        }.items():
            sheet.column_dimensions[column].width = width

        sheet.freeze_panes = "A3"
        sheet.row_dimensions[2].height = 24
        for row in sheet.iter_rows(min_row=3):
            sheet.row_dimensions[row[0].row].height = 28
            for cell in row:
                cell.alignment = text_alignment

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
