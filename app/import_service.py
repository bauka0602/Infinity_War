import base64
import hashlib
import re
from io import BytesIO

from .auth_service import require_auth_user
from .config import DB_LOCK
from .db import db_execute, get_connection, insert_and_get_id, query_all, query_one
from .errors import ApiError
from .lesson_rules import requires_computers_for_component

COURSE_EDUCATIONAL_PROGRAMME_GROUP_ALIASES = {
    "b057": "B057 - Информационные технологии",
    "b057 информационные технологии": "B057 - Информационные технологии",
    "b057 - информационные технологии": "B057 - Информационные технологии",
    "информационные технологии": "B057 - Информационные технологии",
}

IUP_COMPONENT_CODES = {"ОК", "ВК", "КВ", "ДВО", "УПП"}
IUP_LESSON_PHRASES = {
    "Лекции": "lecture",
    "Практики, Семинары": "practical",
    "Лабораторные работы": "lab",
    "Самостоятельная работа студента и преподавателя": "srop",
    "Учебная практика": "practice",
    "Производственная практика": "practice",
}
IUP_ACTIVE_LESSON_TYPES = {"lecture", "practical", "lab"}

ROP_PERIOD_COLUMN_GROUPS = (
    {
        "academic_period_column": 10,
        "total": 10,
        "lecture": 11,
        "practical": 12,
        "lab": 13,
        "studio": 14,
        "practice": 15,
        "srop": 16,
        "sro": 17,
    },
    {
        "academic_period_column": 18,
        "total": 18,
        "lecture": 19,
        "practical": 20,
        "lab": 21,
        "studio": 22,
        "practice": 23,
        "srop": 24,
        "sro": 25,
    },
)

ROP_LESSON_TYPES = ("lecture", "practical", "lab", "studio", "practice", "srop")


def _load_workbook(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel import dependency is not installed on the server.",
        ) from exc

    try:
        return load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ApiError(
            400,
            "bad_request",
            "Не удалось прочитать Excel файл. Используйте формат .xlsx.",
        ) from exc


def _decode_file_payload(payload, allowed_extensions, format_message):
    file_name = (payload.get("fileName") or "").strip()
    file_content = payload.get("fileContent")

    if not file_name or not file_content:
        raise ApiError(
            400,
            "fill_required_fields",
            "Заполните поля: fileName, fileContent",
            {"fields": ["fileName", "fileContent"]},
        )

    if not file_name.lower().endswith(tuple(allowed_extensions)):
        raise ApiError(
            400,
            "bad_request",
            format_message,
        )

    if "," in file_content:
        file_content = file_content.split(",", 1)[1]

    try:
        return file_name, base64.b64decode(file_content)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Некорректное содержимое файла.") from exc


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ").replace("-", "_")


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def _cell_text(value):
    value = _normalize_cell(value)
    return str(value).strip() if value not in (None, "") else ""


def _number_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    normalized = str(value).replace(",", ".").strip()
    if not normalized:
        return None
    try:
        parsed = float(normalized)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def _safe_row_value(row, index):
    return row[index] if index < len(row) else ""


def _normalize_course_department(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return COURSE_EDUCATIONAL_PROGRAMME_GROUP_ALIASES.get(_normalize_header(text), text)

def _upsert_rop_course(connection, course, offering):
    existing = query_one(
        connection,
        """
        SELECT id
        FROM courses
        WHERE lower(code) = lower(?) AND lower(name) = lower(?) AND semester = ? AND year = ? AND lower(programme) = lower(?)
        """,
        (
            course["code"],
            course["name"],
            offering["academicPeriod"],
            course.get("studyYear"),
            course.get("programme") or "",
        ),
    )
    description = (
        f"Imported from ROP. Component: {course.get('component') or '-'}; "
        f"cycle: {course.get('cycle') or '-'}; academic period: {offering['academicPeriod']}."
    )
    update_params = (
        course["name"],
        course["code"],
        course.get("credits"),
        offering.get("totalHours"),
        description,
        course.get("studyYear"),
        offering["academicPeriod"],
        _normalize_course_department(course.get("department") or "B057"),
        course.get("programme") or "",
        course.get("moduleType") or "",
        course.get("moduleName") or "",
        course.get("cycle") or "",
        course.get("component") or "",
        course.get("language") or "",
        course.get("academicYear") or "",
        course.get("entryYear") or "",
        0,
    )
    insert_params = (
        *update_params[:8],
        None,
        "",
        *update_params[8:],
    )

    if existing:
        db_execute(
            connection,
            """
            UPDATE courses
            SET
                name = ?,
                code = ?,
                credits = ?,
                hours = ?,
                description = ?,
                year = ?,
                semester = ?,
                department = ?,
                programme = ?,
                module_type = ?,
                module_name = ?,
                cycle = ?,
                component = ?,
                language = ?,
                academic_year = ?,
                entry_year = ?,
                requires_computers = ?
            WHERE id = ?
            """,
            (*update_params, existing["id"]),
        )
        return "updated", existing["id"]

    course_id = insert_and_get_id(
        connection,
        """
        INSERT INTO courses (
            name, code, credits, hours, description,
            year, semester, department, instructor_id, instructor_name, programme,
            module_type, module_name, cycle, component, language, academic_year, entry_year,
            requires_computers
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        insert_params,
    )
    return "inserted", course_id


def _replace_rop_course_components(connection, course_id, course, offering, lesson_components):
    existing_teachers = {
        (row["lesson_type"], row["academic_period"]): (row.get("teacher_id"), row.get("teacher_name", ""))
        for row in query_all(
            connection,
            """
            SELECT lesson_type, academic_period, teacher_id, teacher_name
            FROM course_components
            WHERE course_id = ?
              AND teacher_id IS NOT NULL
            """,
            (course_id,),
        )
    }
    db_execute(
        connection,
        """
        DELETE FROM course_components
        WHERE course_id = ? AND academic_period = ?
        """,
        (course_id, offering["academicPeriod"]),
    )

    inserted = 0
    for component in lesson_components:
        if (
            component["courseCode"] != course["code"]
            or component["courseName"] != course["name"]
            or component["academicPeriod"] != offering["academicPeriod"]
        ):
            continue

        teacher_id, teacher_name = existing_teachers.get(
            (component["lessonType"], component["academicPeriod"]),
            (None, ""),
        )
        insert_and_get_id(
            connection,
            """
            INSERT INTO course_components (
                course_id, course_code, course_name, programme, study_year,
                academic_period, semester, lesson_type, hours, weekly_classes,
                requires_computers, teacher_id, teacher_name
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                course_id,
                component["courseCode"],
                component["courseName"],
                course.get("programme") or "",
                course.get("studyYear"),
                component["academicPeriod"],
                component["semester"],
                component["lessonType"],
                component["hours"],
                component["weeklyClasses"],
                1 if component.get("requiresComputers") else 0,
                teacher_id,
                teacher_name,
            ),
        )
        inserted += 1

    return inserted


def _load_rop_rows(file_name, file_bytes):
    lower_name = file_name.lower()
    if lower_name.endswith(".xls"):
        try:
            import xlrd
        except ImportError as exc:
            raise ApiError(
                500,
                "internal_server_error",
                "Excel .xls import dependency is not installed on the server.",
            ) from exc

        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
        except Exception as exc:
            raise ApiError(400, "bad_request", "Не удалось прочитать РОП Excel файл.") from exc
        if not workbook.nsheets:
            raise ApiError(400, "bad_request", "В РОП Excel файле нет листов.")
        sheet = workbook.sheet_by_index(0)
        return [
            [_normalize_cell(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel import dependency is not installed on the server.",
        ) from exc

    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Не удалось прочитать РОП Excel файл.") from exc
    sheet = workbook.worksheets[0]
    return [[_normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]


def _extract_quoted_text(text):
    match = re.search(r"[“\"]([^“”\"]+)[”\"]", text)
    return match.group(1).strip() if match else ""


def _extract_academic_year(rows):
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        match = re.search(r"(20\d{2})\s*[-–]\s*(20\d{2})", text)
        if match:
            return f"{match.group(1)}-{match.group(2)}"
    return ""


def _extract_entry_year(rows):
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        if "Год поступления" in text or "Оқуға түскен жылы" in text:
            return text.split(":", 1)[-1].strip()
    return ""


def _extract_programme_name(rows):
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        quoted = _extract_quoted_text(text)
        if quoted and (
            "образовательной программы" in text.lower()
            or "білім беру бағдарламасының" in text.lower()
        ):
            return quoted
    return ""


def _normalize_rop_programme_name(file_name, programme):
    if "сопр" in file_name.lower() and programme:
        return f"{programme} (СОПР)"
    return programme


def _extract_language(file_name, rows):
    lower_name = file_name.lower()
    if "каз" in lower_name or "_kk" in lower_name or "қаз" in lower_name:
        return "kk"
    if "рус" in lower_name or "_ru" in lower_name:
        return "ru"
    for row in rows[:20]:
        text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        if any(marker in text for marker in ("ЖҰМЫС ОҚУ ЖОСПАРЫ", "Оқуға түскен жылы")):
            return "kk"
    return "ru"


def _extract_study_year(file_name, academic_periods):
    match = re.search(r"[_\-\s](\d)(?:[_\-\s]|$)", file_name)
    if match:
        return int(match.group(1))
    for academic_period in academic_periods:
        period_number = _extract_period_number(academic_period)
        if period_number:
            return int((period_number + 1) / 2)
    return None


def _extract_period_number(value):
    match = re.search(r"\d+", _cell_text(value))
    return int(match.group(0)) if match else None


def _semester_in_year(academic_period):
    if not academic_period:
        return None
    return 1 if academic_period % 2 == 1 else 2


def _weekly_classes_from_hours(hours):
    numeric_hours = _number_or_none(hours)
    if not numeric_hours:
        return 0
    return max(1, int(round(float(numeric_hours) / 15)))


def _find_rop_table_header_row(rows):
    for index, row in enumerate(rows):
        normalized_values = {_normalize_header(value) for value in row}
        if (
            "код дисциплины" in normalized_values
            or "пәннің коды" in normalized_values
            or "пәннің_коды" in normalized_values
        ):
            return index
    raise ApiError(400, "bad_request", "В РОП не найдена таблица дисциплин.")


def _is_rop_course_row(row):
    first_cell = _cell_text(_safe_row_value(row, 0)).lower()
    code = _cell_text(_safe_row_value(row, 5))
    name = _cell_text(_safe_row_value(row, 6))
    if not code or not name:
        return False
    if first_cell.startswith(("итого", "средняя", "барлығы", "оқу жоспары", "орташа")):
        return False
    return True


def _decode_iup_payload(payload):
    return _decode_file_payload(
        payload,
        (".pdf", ".xls", ".xlsx"),
        "Поддерживаются PDF или Excel файлы ИУП формата .pdf, .xls, .xlsx.",
    )


def _extract_pdf_text(file_bytes):
    try:
        import fitz
    except ImportError as exc:
        raise ApiError(
            500,
            "iup_pdf_dependency_missing",
            "Для импорта PDF ИУП нужно установить pymupdf.",
        ) from exc

    try:
        document = fitz.open(stream=file_bytes, filetype="pdf")
        return "\n".join(page.get_text("text") for page in document)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Не удалось прочитать PDF ИУП.") from exc


def _extract_iup_excel_text(file_name, file_bytes):
    rows = _load_rop_rows(file_name, file_bytes)
    lines = []
    for row in rows:
        for value in row:
            text = _cell_text(value)
            if text:
                lines.extend(part.strip() for part in text.splitlines() if part.strip())
    return "\n".join(lines)


def _iup_clean_lines(text):
    return [
        line.replace("\xa0", " ").strip()
        for line in text.splitlines()
        if line.replace("\xa0", " ").strip()
    ]


def _join_wrapped_text(lines):
    result = " ".join(lines)
    result = re.sub(r"-\s+", "", result)
    return re.sub(r"\s+", " ", result).strip()


def _normalise_iup_programme(value):
    text = str(value or "").strip()
    if "бизнес" in text.lower():
        return "Бизнес-информатика"
    if "компьютер" in text.lower():
        return "Компьютерная инженерия"
    return text


def _normalise_iup_language(value):
    text = str(value or "").strip().lower()
    if text.startswith(("каз", "қаз")):
        return "kk"
    if text.startswith(("рус", "орыс")):
        return "ru"
    return "ru"


def _infer_iup_group_name(file_name, programme):
    match = re.search(r"(\d{2})[-_](\d{2})", file_name)
    if not match:
        return ""
    base_name = f"05-057-{match.group(1)}-{match.group(2)}"
    if "сопр" in str(programme or "").lower() or "сопр" in file_name.lower():
        return f"{base_name} СОПР"
    return base_name


def _extract_iup_metadata(file_name, lines):
    metadata = {
        "fileName": file_name,
        "groupName": "",
        "programme": "",
        "studyCourse": None,
        "language": "ru",
        "academicYear": "",
    }
    for index, line in enumerate(lines):
        if line.startswith("Курс "):
            match = re.search(r"\d+", line)
            metadata["studyCourse"] = int(match.group(0)) if match else None
        elif line.startswith("Язык обучения"):
            metadata["language"] = _normalise_iup_language(line.replace("Язык обучения", "", 1))
        elif (
            not metadata["academicYear"]
            and "учебный год" in line.lower()
            and re.search(r"20\d{2}\s*[-–]\s*20\d{2}", line)
        ):
            match = re.search(r"(20\d{2})\s*[-–]\s*(20\d{2})", line)
            metadata["academicYear"] = f"{match.group(1)}-{match.group(2)}"
        elif "(6B" in line or "(7M" in line:
            if "информатика" in line.lower() or "инженерия" in line.lower():
                metadata["programme"] = _normalise_iup_programme(re.sub(r"\s*\([^)]*\)", "", line))

    metadata["groupName"] = _infer_iup_group_name(file_name, metadata["programme"])
    return metadata


def _iup_lesson_at(lines, index):
    for size in range(1, 4):
        phrase = " ".join(lines[index : index + size])
        lesson_type = IUP_LESSON_PHRASES.get(phrase)
        if lesson_type:
            return lesson_type, size, phrase
    return None, 0, ""


def _is_iup_course_start(lines, index):
    return (
        index + 1 < len(lines)
        and re.fullmatch(r"\d{1,3}", lines[index] or "")
        and lines[index + 1] in IUP_COMPONENT_CODES
    )


def _split_iup_code_and_name(values):
    code_lines = []
    cursor = 0
    while cursor < len(values):
        line = values[cursor]
        if re.search(r"\d", line):
            code_lines.append(line)
            cursor += 1
            continue
        if (
            not code_lines
            and re.fullmatch(r"[A-Za-z]{2,12}", line)
            and cursor + 1 < len(values)
            and re.fullmatch(r"\d{4}", values[cursor + 1])
        ):
            code_lines.append(line)
            cursor += 1
            continue
        break
    return _join_wrapped_text(code_lines), _join_wrapped_text(values[cursor:])


def _parse_iup_course_block(block):
    if len(block) < 5:
        return None
    course_index = int(block[0])
    component = block[1]
    credits_index = None
    for index in range(2, len(block) - 1):
        if re.fullmatch(r"\d{1,2}", block[index]) and _iup_lesson_at(block, index + 1)[0]:
            credits_index = index
            break
    if credits_index is None:
        return None

    course_code, course_name = _split_iup_code_and_name(block[2:credits_index])
    if not course_code or not course_name:
        return None

    lesson_items = []
    cursor = credits_index + 1
    while cursor < len(block):
        lesson_type, size, phrase = _iup_lesson_at(block, cursor)
        if not lesson_type:
            cursor += 1
            continue
        cursor += size
        teacher_lines = []
        hours = None
        while cursor < len(block):
            hours_match = re.match(r"^(\d+)\b", block[cursor])
            if hours_match:
                hours = int(hours_match.group(1))
                cursor += 1
                break
            if _iup_lesson_at(block, cursor)[0]:
                break
            teacher_lines.append(block[cursor])
            cursor += 1

        teacher_name = _join_wrapped_text(teacher_lines)
        if teacher_name and hours is not None:
            lesson_items.append(
                {
                    "lessonType": lesson_type,
                    "lessonLabel": phrase,
                    "teacherName": teacher_name,
                    "hours": hours,
                }
            )

    return {
        "index": course_index,
        "component": component,
        "code": course_code,
        "name": course_name,
        "credits": int(block[credits_index]),
        "lessonItems": lesson_items,
    }


def _parse_iup_file(file_name, file_bytes):
    lower_name = file_name.lower()
    if lower_name.endswith(".pdf"):
        raw_text = _extract_pdf_text(file_bytes)
    elif lower_name.endswith((".xls", ".xlsx")):
        raw_text = _extract_iup_excel_text(file_name, file_bytes)
    else:
        raise ApiError(400, "bad_request", "Поддерживаются PDF или Excel файлы ИУП.")

    lines = _iup_clean_lines(raw_text)
    metadata = _extract_iup_metadata(file_name, lines)
    courses = []
    entries = []
    current_study_year = metadata.get("studyCourse")
    current_academic_period = None
    index = 0

    while index < len(lines):
        course_year_match = re.match(r"^(\d+)\s+Курс обучения", lines[index])
        if course_year_match:
            current_study_year = int(course_year_match.group(1))

        period_match = re.match(r"^(\d+)\s+Академический период", lines[index])
        if period_match:
            current_academic_period = int(period_match.group(1))

        if not _is_iup_course_start(lines, index):
            index += 1
            continue

        next_index = index + 2
        while next_index < len(lines) and not _is_iup_course_start(lines, next_index):
            next_index += 1

        course = _parse_iup_course_block(lines[index:next_index])
        if course:
            course["studyYear"] = current_study_year
            course["academicPeriod"] = current_academic_period
            course["semester"] = current_academic_period
            courses.append(course)
            for lesson in course["lessonItems"]:
                entries.append(
                    {
                        **metadata,
                        "studyYear": current_study_year,
                        "academicPeriod": current_academic_period,
                        "semester": current_academic_period,
                        "component": course["component"],
                        "courseCode": course["code"],
                        "courseName": course["name"],
                        "credits": course["credits"],
                        **lesson,
                    }
                )

        index = next_index

    if not courses:
        raise ApiError(400, "bad_request", "В ИУП не найдены дисциплины.")

    return {
        "metadata": metadata,
        "courses": courses,
        "entries": entries,
        "totals": {
            "courses": len(courses),
            "lessonEntries": len(entries),
            "teachers": len({entry["teacherName"] for entry in entries if entry.get("teacherName")}),
        },
    }


def _teacher_email_from_name(name):
    digest = hashlib.sha1(name.strip().lower().encode("utf-8")).hexdigest()[:12]
    return f"iup-{digest}@imported.local"


def _upsert_iup_teacher(connection, teacher_name, language):
    existing = query_one(
        connection,
        "SELECT id, name FROM teachers WHERE lower(name) = lower(?)",
        (teacher_name,),
    )
    if existing:
        return existing["id"], "existing"

    teacher_id = insert_and_get_id(
        connection,
        """
        INSERT INTO teachers (name, email, department, teaching_languages)
        VALUES (?, ?, ?, ?)
        """,
        (
            teacher_name,
            _teacher_email_from_name(teacher_name),
            "Институт бизнеса и цифровых технологий",
            language or "ru,kk",
        ),
    )
    return teacher_id, "inserted"


def _resolve_iup_group_name(connection, group_name):
    if not group_name:
        return ""
    exact = query_one(connection, "SELECT name FROM groups WHERE name = ?", (group_name,))
    if exact:
        return exact["name"]
    prefixed = query_one(
        connection,
        "SELECT name FROM groups WHERE name LIKE ? ORDER BY length(name), name LIMIT 1",
        (f"{group_name}%",),
    )
    return prefixed["name"] if prefixed else group_name


def _find_matching_iup_course(connection, course_code, programme, semester):
    return query_one(
        connection,
        """
        SELECT id, instructor_id
        FROM courses
        WHERE lower(code) = lower(?)
          AND (
            ? = ''
            OR lower(programme) = lower(?)
            OR lower(programme) LIKE lower(?)
            OR programme = ''
          )
          AND (? IS NULL OR semester = ?)
        ORDER BY
          CASE
            WHEN lower(programme) = lower(?) THEN 0
            WHEN lower(programme) LIKE lower(?) THEN 1
            WHEN programme = '' THEN 2
            ELSE 3
          END,
          id
        LIMIT 1
        """,
        (
            course_code,
            programme or "",
            programme or "",
            f"%{programme}%" if programme else "",
            semester,
            semester,
            programme or "",
            f"%{programme}%" if programme else "",
        ),
    )


def _find_matching_iup_course_relaxed(connection, course_code, programme, semester):
    course = _find_matching_iup_course(connection, course_code, programme, semester)
    if course:
        return course

    if semester is not None:
        course = query_one(
            connection,
            """
            SELECT id, instructor_id
            FROM courses
            WHERE lower(code) = lower(?)
              AND semester = ?
            ORDER BY id
            LIMIT 1
            """,
            (course_code, semester),
        )
        if course:
            return course

    return query_one(
        connection,
        """
        SELECT id, instructor_id
        FROM courses
        WHERE lower(code) = lower(?)
        ORDER BY id
        LIMIT 1
        """,
        (course_code,),
    )


def _course_import_item(code, name, semester=None, programme=None):
    return {
        "code": code or "",
        "name": name or "",
        "semester": semester,
        "programme": programme or "",
    }


def _get_iup_course_lists(connection, parsed):
    metadata = parsed["metadata"]
    matched_courses = []
    missing_courses = []
    seen_iup_courses = set()

    for course in parsed["courses"]:
        key = (
            (course.get("code") or "").lower(),
            (course.get("name") or "").lower(),
            course.get("semester"),
        )
        if key in seen_iup_courses:
            continue
        seen_iup_courses.add(key)
        match = _find_matching_iup_course_relaxed(
            connection,
            course.get("code", ""),
            metadata.get("programme", ""),
            course.get("semester"),
        )
        item = _course_import_item(
            course.get("code", ""),
            course.get("name", ""),
            course.get("semester"),
            metadata.get("programme", ""),
        )
        if match:
            matched_courses.append(item)
        else:
            missing_courses.append(item)

    return matched_courses, missing_courses


def _create_iup_missing_courses(connection, parsed, missing_courses):
    metadata = parsed["metadata"]
    created = []
    courses_by_key = {
        ((course.get("code") or "").lower(), course.get("semester")): course
        for course in parsed["courses"]
    }

    for item in missing_courses:
        course = courses_by_key.get(((item.get("code") or "").lower(), item.get("semester"))) or item
        course_id = insert_and_get_id(
            connection,
            """
            INSERT INTO courses (
                name, code, credits, hours, description,
                year, semester, department, instructor_id, instructor_name, programme,
                module_type, module_name, cycle, component, language, academic_year, entry_year,
                requires_computers
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                course.get("name") or item.get("name") or item.get("code") or "Без названия",
                course.get("code") or item.get("code") or "",
                course.get("credits"),
                None,
                "Imported from IUP as missing course draft.",
                course.get("studyYear"),
                course.get("semester") or item.get("semester"),
                _normalize_course_department(metadata.get("educationalProgrammeGroup") or "B057"),
                None,
                "",
                metadata.get("programme", ""),
                "",
                "",
                "",
                course.get("component", ""),
                metadata.get("language", "ru"),
                metadata.get("academicYear", ""),
                "",
                0,
            ),
        )
        created.append(_course_import_item(
            course.get("code") or item.get("code"),
            course.get("name") or item.get("name"),
            course.get("semester") or item.get("semester"),
            metadata.get("programme", ""),
        ))

        lesson_entries = [
            entry
            for entry in parsed["entries"]
            if (entry.get("courseCode") or "").lower() == (course.get("code") or item.get("code") or "").lower()
               and entry.get("lessonType") in IUP_ACTIVE_LESSON_TYPES
        ]
        seen_components = set()
        for entry in lesson_entries:
            component_key = (entry.get("lessonType"), entry.get("academicPeriod"))
            if component_key in seen_components:
                continue
            seen_components.add(component_key)
            insert_and_get_id(
                connection,
                """
                INSERT INTO course_components (
                    course_id, course_code, course_name, lesson_type, hours,
                    weekly_classes, academic_period, semester, requires_computers,
                    teacher_id, teacher_name
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    course_id,
                    entry.get("courseCode", ""),
                    entry.get("courseName", ""),
                    entry.get("lessonType"),
                    entry.get("hours"),
                    1,
                    entry.get("academicPeriod"),
                    entry.get("semester"),
                    1 if requires_computers_for_component(
                        entry.get("lessonType"),
                        entry.get("courseCode", ""),
                        entry.get("courseName", ""),
                        entry.get("studyYear"),
                    ) else 0,
                    None,
                    "",
                ),
            )

    return created


def _store_iup_entries(connection, parsed, create_missing_courses=False):
    metadata = parsed["metadata"]
    metadata["groupName"] = _resolve_iup_group_name(connection, metadata.get("groupName", ""))
    db_execute(
        connection,
        "DELETE FROM iup_entries WHERE file_name = ? AND group_name = ?",
        (metadata["fileName"], metadata.get("groupName", "")),
    )

    updated_courses = set()
    updated_components = set()
    teacher_cache = {}
    matched_courses, missing_courses = _get_iup_course_lists(connection, parsed)
    created_courses = []
    if create_missing_courses and missing_courses:
        created_courses = _create_iup_missing_courses(connection, parsed, missing_courses)
        matched_courses, missing_courses = _get_iup_course_lists(connection, parsed)

    for entry in parsed["entries"]:
        teacher_name = entry.get("teacherName", "")
        teacher_id = None
        if teacher_name:
            if teacher_name not in teacher_cache:
                teacher_cache[teacher_name] = _upsert_iup_teacher(
                    connection,
                    teacher_name,
                    metadata.get("language", "ru"),
                )
            teacher_id, status = teacher_cache[teacher_name]

        insert_and_get_id(
            connection,
            """
            INSERT INTO iup_entries (
                file_name, group_name, programme, study_course,
                language, academic_year, academic_period, semester, component,
                course_code, course_name, credits, lesson_type, teacher_id,
                teacher_name, hours
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                metadata["fileName"],
                metadata.get("groupName", ""),
                metadata.get("programme", ""),
                entry.get("studyYear"),
                metadata.get("language", "ru"),
                metadata.get("academicYear", ""),
                entry.get("academicPeriod"),
                entry.get("semester"),
                entry.get("component", ""),
                entry["courseCode"],
                entry["courseName"],
                entry.get("credits"),
                entry["lessonType"],
                teacher_id,
                teacher_name,
                entry.get("hours"),
            ),
        )

        if teacher_id and entry["lessonType"] in IUP_ACTIVE_LESSON_TYPES:
            course = _find_matching_iup_course_relaxed(
                connection,
                entry["courseCode"],
                metadata.get("programme", ""),
                entry.get("semester"),
            )
            if course and course["id"] not in updated_courses:
                db_execute(
                    connection,
                    """
                    UPDATE courses
                    SET instructor_id = ?, instructor_name = ?
                    WHERE id = ?
                    """,
                    (teacher_id, teacher_name, course["id"]),
                )
                updated_courses.add(course["id"])

            if course:
                component_update = db_execute(
                    connection,
                    """
                    UPDATE course_components
                    SET teacher_id = ?, teacher_name = ?
                    WHERE course_id = ?
                      AND lesson_type = ?
                      AND (? IS NULL OR academic_period = ?)
                    """,
                    (
                        teacher_id,
                        teacher_name,
                        course["id"],
                        entry["lessonType"],
                        entry.get("academicPeriod"),
                        entry.get("academicPeriod"),
                    ),
                )
                updated_count = max(0, getattr(component_update, "rowcount", 0) or 0)
                if updated_count == 0 and entry.get("academicPeriod") is not None:
                    fallback_update = db_execute(
                        connection,
                        """
                        UPDATE course_components
                        SET teacher_id = ?, teacher_name = ?
                        WHERE course_id = ?
                          AND lesson_type = ?
                        """,
                        (
                            teacher_id,
                            teacher_name,
                            course["id"],
                            entry["lessonType"],
                        ),
                    )
                    updated_count = max(0, getattr(fallback_update, "rowcount", 0) or 0)
                db_execute(
                    connection,
                    """
                    UPDATE sections
                    SET teacher_id = ?, teacher_name = ?
                    WHERE course_id = ?
                      AND lesson_type = ?
                    """,
                    (
                        teacher_id,
                        teacher_name,
                        course["id"],
                        entry["lessonType"],
                    ),
                )
                if updated_count:
                    updated_components.add((entry["courseCode"], entry["lessonType"], entry.get("academicPeriod")))

    return {
        "iupEntries": len(parsed["entries"]),
        "teachersInserted": sum(1 for _teacher_id, status in teacher_cache.values() if status == "inserted"),
        "teachersExisting": sum(1 for _teacher_id, status in teacher_cache.values() if status == "existing"),
        "coursesUpdated": len(updated_courses),
        "componentsUpdated": len(updated_components),
        "coursesCreated": len(created_courses),
        "coursesExisting": len(matched_courses),
        "coursesMissing": len(missing_courses),
        "courseLists": {
            "inserted": created_courses,
            "existing": matched_courses,
            "missing": missing_courses,
        },
    }


def parse_iup_preview(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    file_name, file_bytes = _decode_iup_payload(payload)
    parsed = _parse_iup_file(file_name, file_bytes)
    with DB_LOCK:
        with get_connection() as connection:
            matched_courses, missing_courses = _get_iup_course_lists(connection, parsed)
    return {
        "metadata": parsed["metadata"],
        "totals": parsed["totals"],
        "courses": [
            {
                "code": course["code"],
                "name": course["name"],
                "component": course["component"],
                "credits": course["credits"],
                "academicPeriod": course.get("academicPeriod"),
            }
            for course in parsed["courses"]
        ],
        "entries": parsed["entries"][:50],
        "courseLists": {
            "existing": matched_courses,
            "missing": missing_courses,
        },
    }


def import_iup_data(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    file_name, file_bytes = _decode_iup_payload(payload)
    parsed = _parse_iup_file(file_name, file_bytes)
    create_missing_courses = bool(payload.get("createMissingCourses"))
    with DB_LOCK:
        with get_connection() as connection:
            stats = _store_iup_entries(connection, parsed, create_missing_courses)
            connection.commit()
    return {
        "success": True,
        "metadata": parsed["metadata"],
        "totals": parsed["totals"],
        "stats": stats,
    }


def parse_rop_preview(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    file_name, file_bytes = _decode_file_payload(
        payload,
        (".xls", ".xlsx"),
        "Поддерживаются Excel файлы РОП формата .xls или .xlsx.",
    )
    rows = _load_rop_rows(file_name, file_bytes)
    header_row_index = _find_rop_table_header_row(rows)
    academic_period_row = rows[header_row_index + 1] if header_row_index + 1 < len(rows) else []
    academic_periods = [
        _extract_period_number(_safe_row_value(academic_period_row, group["academic_period_column"]))
        for group in ROP_PERIOD_COLUMN_GROUPS
    ]

    study_year = _extract_study_year(file_name, academic_periods)
    metadata = {
        "fileName": file_name,
        "programme": _normalize_rop_programme_name(file_name, _extract_programme_name(rows)),
        "academicYear": _extract_academic_year(rows),
        "entryYear": _extract_entry_year(rows),
        "language": _extract_language(file_name, rows),
        "studyYear": study_year,
        "academicPeriods": [period for period in academic_periods if period],
    }

    courses = []
    offerings = []
    lesson_components = []
    seen_courses = set()

    for row_index, row in enumerate(rows[header_row_index + 3 :], start=header_row_index + 4):
        if not _is_rop_course_row(row):
            continue

        course = {
            "rowNumber": row_index,
            "moduleNumber": _cell_text(_safe_row_value(row, 0)),
            "moduleType": _cell_text(_safe_row_value(row, 1)),
            "moduleName": _cell_text(_safe_row_value(row, 2)),
            "cycle": _cell_text(_safe_row_value(row, 3)),
            "component": _cell_text(_safe_row_value(row, 4)),
            "code": _cell_text(_safe_row_value(row, 5)),
            "name": _cell_text(_safe_row_value(row, 6)),
            "credits": _number_or_none(_safe_row_value(row, 7)),
            "examPeriod": _number_or_none(_safe_row_value(row, 9)),
            "programme": metadata["programme"],
            "studyYear": study_year,
            "language": metadata["language"],
            "academicYear": metadata["academicYear"],
            "entryYear": metadata["entryYear"],
        }
        course_key = (course["code"].lower(), course["name"].lower())
        if course_key not in seen_courses:
            seen_courses.add(course_key)
            courses.append(course)

        for group, academic_period in zip(ROP_PERIOD_COLUMN_GROUPS, academic_periods):
            total_hours = _number_or_none(_safe_row_value(row, group["total"]))
            if not academic_period or not total_hours:
                continue

            offering = {
                "courseCode": course["code"],
                "courseName": course["name"],
                "academicPeriod": academic_period,
                "semester": _semester_in_year(academic_period),
                "studyYear": study_year,
                "totalHours": total_hours,
                "lectureHours": _number_or_none(_safe_row_value(row, group["lecture"])) or 0,
                "practicalHours": _number_or_none(_safe_row_value(row, group["practical"])) or 0,
                "labHours": _number_or_none(_safe_row_value(row, group["lab"])) or 0,
                "studioHours": _number_or_none(_safe_row_value(row, group["studio"])) or 0,
                "practiceHours": _number_or_none(_safe_row_value(row, group["practice"])) or 0,
                "sropHours": _number_or_none(_safe_row_value(row, group["srop"])) or 0,
                "sroHours": _number_or_none(_safe_row_value(row, group["sro"])) or 0,
            }
            offerings.append(offering)

            for lesson_type in ROP_LESSON_TYPES:
                hours = offering[f"{lesson_type}Hours"]
                if hours:
                    normalized_lesson_type = "practical" if lesson_type == "studio" else lesson_type
                    lesson_components.append(
                        {
                            "courseCode": course["code"],
                            "courseName": course["name"],
                            "programme": metadata["programme"],
                            "studyYear": study_year,
                            "academicPeriod": academic_period,
                            "semester": offering["semester"],
                            "lessonType": normalized_lesson_type,
                            "hours": hours,
                            "weeklyClasses": _weekly_classes_from_hours(hours),
                            "requiresComputers": requires_computers_for_component(
                                normalized_lesson_type,
                                course["code"],
                                course["name"],
                                study_year,
                            ),
                        }
                    )

    if not courses:
        raise ApiError(400, "bad_request", "В РОП не найдены дисциплины.")

    return {
        "message": "ROP preview parsed successfully.",
        "metadata": metadata,
        "totals": {
            "courses": len(courses),
            "offerings": len(offerings),
            "lessonComponents": len(lesson_components),
        },
        "courses": courses,
        "offerings": offerings,
        "lessonComponents": lesson_components,
    }


def import_rop_data(headers, payload):
    preview = parse_rop_preview(headers, payload)
    course_by_key = {
        (course["code"], course["name"]): course for course in preview["courses"]
    }
    summary = {
        "courses": {"inserted": 0, "updated": 0},
        "courseComponents": {"inserted": 0},
    }
    course_lists = {
        "inserted": [],
        "existing": [],
    }
    seen_course_results = set()

    with DB_LOCK:
        with get_connection() as connection:
            for offering in preview["offerings"]:
                course = course_by_key.get((offering["courseCode"], offering["courseName"]))
                if not course:
                    continue
                result, course_id = _upsert_rop_course(connection, course, offering)
                summary["courses"][result] += 1
                list_key = (
                    result,
                    course["code"].lower(),
                    course["name"].lower(),
                    offering["academicPeriod"],
                    (course.get("programme") or "").lower(),
                )
                if list_key not in seen_course_results:
                    seen_course_results.add(list_key)
                    course_lists["inserted" if result == "inserted" else "existing"].append(
                        _course_import_item(
                            course["code"],
                            course["name"],
                            offering["academicPeriod"],
                            course.get("programme") or "",
                        )
                    )
                summary["courseComponents"]["inserted"] += _replace_rop_course_components(
                    connection,
                    course_id,
                    course,
                    offering,
                    preview["lessonComponents"],
                )
            connection.commit()

    return {
        "message": "ROP import completed successfully.",
        "metadata": preview["metadata"],
        "summary": summary,
        "courseLists": course_lists,
        "totals": {
            "inserted": summary["courses"]["inserted"],
            "updated": summary["courses"]["updated"],
            "courseComponents": summary["courseComponents"]["inserted"],
            "courses": preview["totals"]["courses"],
            "offerings": preview["totals"]["offerings"],
            "lessonComponents": preview["totals"]["lessonComponents"],
        },
    }


def generate_schedule_export(headers):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel export dependency is not installed on the server.",
        ) from exc

    with DB_LOCK:
        with get_connection() as connection:
            schedules = query_all(
                connection,
                """
                SELECT
                    course_name,
                    group_name,
                    subgroup,
                    teacher_name,
                    room_number,
                    day,
                    start_hour,
                    semester,
                    year,
                    algorithm
                FROM schedules
                ORDER BY day, start_hour, course_name, group_name, id
                """,
            )

    if not schedules:
        raise ApiError(400, "bad_request", "Расписание ещё не сгенерировано.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"
    sheet.append(
        [
            "course_name",
            "group_name",
            "subgroup",
            "teacher_name",
            "room_number",
            "day",
            "start_hour",
            "semester",
            "year",
            "algorithm",
        ]
    )

    for item in schedules:
        sheet.append(
            [
                item.get("course_name", ""),
                item.get("group_name", ""),
                item.get("subgroup", ""),
                item.get("teacher_name", ""),
                item.get("room_number", ""),
                item.get("day", ""),
                item.get("start_hour", ""),
                item.get("semester", ""),
                item.get("year", ""),
                item.get("algorithm", ""),
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
