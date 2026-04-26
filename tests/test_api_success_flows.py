import base64
import time
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from backend.app.lesson_rules import requires_computers_for_component


def _wait_for_job_completion(client, headers, job_id, timeout_seconds=5):
    started_at = time.time()
    while time.time() - started_at < timeout_seconds:
        response = client.get(f"/api/schedules/generate/{job_id}", headers=headers)
        assert response.status_code == 200
        payload = response.json()
        if payload["status"] in {"completed", "failed"}:
            return payload
        time.sleep(0.1)
    raise AssertionError(f"Job {job_id} did not finish within {timeout_seconds} seconds")


def _build_rop_preview_payload():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"

    def row_with(values):
        row = [""] * 26
        for index, value in values.items():
            row[index] = value
        sheet.append(row)

    row_with({0: 'РАБОЧИЙ УЧЕБНЫЙ ПЛАН на 2025-2026 учебные годы'})
    row_with({0: 'для Модульной образовательной программы “Бизнес-информатика”'})
    row_with({0: 'Год поступления: 01-09-2024'})
    row_with({})
    row_with(
        {
            0: "№ модуля",
            1: "Тип модуля",
            2: "Наименование модуля",
            3: "Цикл дисциплины",
            4: "Компонент дисциплины",
            5: "Код дисциплины",
            6: "Наименование дисциплины",
            7: "Академические кредиты",
            9: "Экзамены (семестр)*",
            10: "Распределение часов",
        }
    )
    row_with({10: "3 Академический период", 18: "4 Академический период"})
    row_with(
        {
            10: "Всего",
            11: "Лекции",
            12: "Практические",
            13: "Лабораторные",
            18: "Всего",
            19: "Лекции",
            20: "Практические",
            21: "Лабораторные",
        }
    )
    row_with(
        {
            0: 1,
            1: "Общие модули",
            2: "Гуманитарно-социальный",
            3: "ООД",
            4: "ОК",
            5: "Fil 2108",
            6: "Философия",
            7: 5,
            9: 1,
            10: 150,
            11: 15,
            12: 30,
            15: 30,
            16: 20,
            17: 85,
        }
    )

    buffer = BytesIO()
    workbook.save(buffer)
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return {"fileName": "БИ_2_рус.xlsx", "fileContent": encoded}


def _seed_schedule_data(client, headers):
    teacher_response = client.post(
        "/api/teachers",
        headers=headers,
        json={
            "name": "Aruzhan Saparova",
            "email": "aruzhan.saparova@kazatu.edu.kz",
            "phone": "+7 777 000 00 00",
            "department": "CS",
            "teaching_languages": "ru,kk",
        },
    )
    assert teacher_response.status_code == 201
    teacher = teacher_response.json()

    course_response = client.post(
        "/api/disciplines",
        headers=headers,
        json={
            "code": "CS201",
            "name": "Algorithms",
            "credits": 5,
            "hours": 150,
            "year": 2,
            "semester": 1,
            "programme": "Software Engineering",
            "department": "B057 - Информационные технологии",
            "instructor_id": teacher["id"],
            "instructor_name": teacher["name"],
        },
    )
    assert course_response.status_code == 201
    course = course_response.json()

    room_response = client.post(
        "/api/rooms",
        headers=headers,
        json={
            "number": "401",
            "capacity": 40,
            "building": "Main",
            "type": "lecture",
            "department": "CS",
            "available": 1,
        },
    )
    assert room_response.status_code == 201

    group_response = client.post(
        "/api/groups",
        headers=headers,
        json={
            "name": "SE-24-01",
            "student_count": 24,
            "study_course": 2,
            "has_subgroups": 0,
            "language": "ru",
            "programme": "Software Engineering",
        },
    )
    assert group_response.status_code == 201
    group = group_response.json()

    section_response = client.post(
        "/api/sections",
        headers=headers,
        json={
            "course_id": course["id"],
            "course_name": course["name"],
            "group_id": group["id"],
            "group_name": group["name"],
            "classes_count": 2,
            "lesson_type": "lecture",
        },
    )
    assert section_response.status_code == 201


def test_rop_preview_parses_curriculum_plan(client, admin_auth_headers):
    response = client.post(
        "/api/import/rop/preview",
        headers=admin_auth_headers,
        json=_build_rop_preview_payload(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["metadata"]["programme"] == "Бизнес-информатика"
    assert payload["metadata"]["academicYear"] == "2025-2026"
    assert payload["metadata"]["entryYear"] == "01-09-2024"
    assert payload["metadata"]["language"] == "ru"
    assert payload["metadata"]["studyYear"] == 2
    assert payload["metadata"]["academicPeriods"] == [3, 4]
    assert payload["totals"]["courses"] == 1
    assert payload["totals"]["offerings"] == 1
    assert payload["totals"]["lessonComponents"] == 4
    assert payload["courses"][0]["code"] == "Fil 2108"
    assert payload["offerings"][0]["semester"] == 1
    assert {item["lessonType"] for item in payload["lessonComponents"]} == {
        "lecture",
        "practical",
        "practice",
        "srop",
    }
    computer_flags = {
        item["lessonType"]: item["requiresComputers"]
        for item in payload["lessonComponents"]
    }
    assert computer_flags == {
        "lecture": False,
        "practical": False,
        "practice": False,
        "srop": False,
    }


def test_rop_import_creates_courses_from_curriculum_plan(client, admin_auth_headers):
    response = client.post(
        "/api/import/rop",
        headers=admin_auth_headers,
        json=_build_rop_preview_payload(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["courses"] == {"inserted": 1, "updated": 0}

    courses_response = client.get("/api/disciplines", headers=admin_auth_headers)
    assert courses_response.status_code == 200
    courses = courses_response.json()
    assert len(courses) == 1
    assert courses[0]["code"] == "Fil 2108"
    assert courses[0]["name"] == "Философия"
    assert courses[0]["credits"] == 5
    assert courses[0]["hours"] == 150
    assert courses[0]["year"] == 2
    assert courses[0]["semester"] == 3
    assert courses[0]["programme"] == "Бизнес-информатика"
    assert courses[0]["cycle"] == "ООД"
    assert courses[0]["component"] == "ОК"
    assert courses[0]["academic_year"] == "2025-2026"
    assert courses[0]["entry_year"] == "01-09-2024"

    components_response = client.get("/api/course_components", headers=admin_auth_headers)
    assert components_response.status_code == 200
    components = components_response.json()
    assert len(components) == 4
    assert {item["lesson_type"] for item in components} == {"lecture", "practical", "practice", "srop"}
    assert {item["requires_computers"] for item in components} == {0}


def test_practical_requires_computers_only_for_it_after_first_year():
    assert requires_computers_for_component("lab", "Lang 1101", "Английский язык", 1) is True
    assert requires_computers_for_component("practical", "IKT 1101", "Информационные технологии", 1) is False
    assert requires_computers_for_component("practical", "BD 2201", "Базы данных", 2) is True
    assert requires_computers_for_component("practice", "Pr 4301", "Преддипломная практика", 4) is False


def test_rop_import_preserves_existing_course_instructor(client, admin_auth_headers):
    teacher_response = client.post(
        "/api/teachers",
        headers=admin_auth_headers,
        json={
            "name": "Aruzhan Saparova",
            "email": "aruzhan.saparova@kazatu.edu.kz",
            "phone": "+7 777 000 00 00",
            "department": "B057 - Информационные технологии",
            "teaching_languages": "ru,kk",
        },
    )
    assert teacher_response.status_code == 201
    teacher = teacher_response.json()

    course_response = client.post(
        "/api/disciplines",
        headers=admin_auth_headers,
        json={
            "code": "Fil 2108",
            "name": "Философия",
            "credits": 4,
            "hours": 120,
            "year": 2,
            "semester": 3,
            "department": "B057 - Информационные технологии",
            "programme": "Бизнес-информатика",
            "instructor_id": teacher["id"],
            "instructor_name": teacher["name"],
        },
    )
    assert course_response.status_code == 201

    response = client.post(
        "/api/import/rop",
        headers=admin_auth_headers,
        json=_build_rop_preview_payload(),
    )
    assert response.status_code == 200
    assert response.json()["summary"]["courses"] == {"inserted": 0, "updated": 1}

    courses_response = client.get("/api/disciplines", headers=admin_auth_headers)
    assert courses_response.status_code == 200
    [course] = courses_response.json()
    assert course["credits"] == 5
    assert course["hours"] == 150
    assert course["instructor_id"] == teacher["id"]
    assert course["instructor_name"] == teacher["name"]


def test_section_uses_teacher_from_matching_course_component(client, admin_auth_headers, backend_modules):
    lecture_teacher_response = client.post(
        "/api/teachers",
        headers=admin_auth_headers,
        json={
            "name": "Lecture Teacher",
            "email": "lecture.teacher@kazatu.edu.kz",
            "department": "B057 - Информационные технологии",
            "teaching_languages": "ru,kk",
        },
    )
    assert lecture_teacher_response.status_code == 201
    lecture_teacher = lecture_teacher_response.json()

    practical_teacher_response = client.post(
        "/api/teachers",
        headers=admin_auth_headers,
        json={
            "name": "Practical Teacher",
            "email": "practical.teacher@kazatu.edu.kz",
            "department": "B057 - Информационные технологии",
            "teaching_languages": "ru,kk",
        },
    )
    assert practical_teacher_response.status_code == 201
    practical_teacher = practical_teacher_response.json()

    course_response = client.post(
        "/api/disciplines",
        headers=admin_auth_headers,
        json={
            "code": "CS301",
            "name": "Databases",
            "credits": 5,
            "hours": 150,
            "year": 2,
            "semester": 1,
            "department": "B057 - Информационные технологии",
            "programme": "Бизнес-информатика",
            "instructor_id": lecture_teacher["id"],
            "instructor_name": lecture_teacher["name"],
        },
    )
    assert course_response.status_code == 201
    course = course_response.json()

    _app_module, db_module = backend_modules
    with db_module.get_connection() as connection:
        db_module.insert_and_get_id(
            connection,
            """
            INSERT INTO course_components (
                course_id, course_code, course_name, programme, study_year,
                academic_period, semester, lesson_type, hours, weekly_classes,
                requires_computers, teacher_id, teacher_name
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                course["id"],
                course["code"],
                course["name"],
                course["programme"],
                course["year"],
                course["semester"],
                course["semester"],
                "practical",
                30,
                2,
                1,
                practical_teacher["id"],
                practical_teacher["name"],
            ),
        )
        connection.commit()

    group_response = client.post(
        "/api/groups",
        headers=admin_auth_headers,
        json={
            "name": "BI-24-01",
            "student_count": 24,
            "language": "ru",
            "study_course": 2,
        },
    )
    assert group_response.status_code == 201
    group = group_response.json()

    section_response = client.post(
        "/api/sections",
        headers=admin_auth_headers,
        json={
            "course_id": course["id"],
            "course_name": course["name"],
            "group_id": group["id"],
            "group_name": group["name"],
            "classes_count": 1,
            "lesson_type": "practical",
        },
    )
    assert section_response.status_code == 201
    section = section_response.json()
    assert section["teacher_id"] == practical_teacher["id"]
    assert section["teacher_name"] == practical_teacher["name"]


def test_generate_sections_from_rop_components_for_matching_groups(client, admin_auth_headers):
    rop_response = client.post(
        "/api/import/rop",
        headers=admin_auth_headers,
        json=_build_rop_preview_payload(),
    )
    assert rop_response.status_code == 200

    group_response = client.post(
        "/api/groups",
        headers=admin_auth_headers,
        json={
            "name": "BI-24-01",
            "student_count": 24,
            "language": "ru",
            "programme": "Бизнес-информатика",
            "study_course": 2,
        },
    )
    assert group_response.status_code == 201

    response = client.post(
        "/api/sections/generate",
        headers=admin_auth_headers,
        json={
            "programme": "Бизнес-информатика",
            "study_course": 2,
            "semester": 3,
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["inserted"] == 2
    assert payload["updated"] == 0
    assert {section["lesson_type"] for section in payload["sections"]} == {"lecture", "practical"}

    second_response = client.post(
        "/api/sections/generate",
        headers=admin_auth_headers,
        json={
            "programme": "Бизнес-информатика",
            "study_course": 2,
            "semester": 3,
        },
    )
    assert second_response.status_code == 200
    assert second_response.json()["inserted"] == 0
    assert second_response.json()["updated"] == 2


def test_generate_sections_without_filters_uses_all_matching_data(client, admin_auth_headers):
    rop_response = client.post(
        "/api/import/rop",
        headers=admin_auth_headers,
        json=_build_rop_preview_payload(),
    )
    assert rop_response.status_code == 200

    for group_name in ["BI-24-01", "BI-24-02"]:
        group_response = client.post(
            "/api/groups",
            headers=admin_auth_headers,
            json={
                "name": group_name,
                "student_count": 24,
                "language": "ru",
                "programme": "Бизнес-информатика",
                "study_course": 2,
            },
        )
        assert group_response.status_code == 201

    response = client.post("/api/sections/generate", headers=admin_auth_headers, json={})
    assert response.status_code == 200
    payload = response.json()
    assert payload["inserted"] == 4
    assert payload["updated"] == 0
    assert {
        (section["group_name"], section["lesson_type"])
        for section in payload["sections"]
    } == {
        ("BI-24-01", "lecture"),
        ("BI-24-01", "practical"),
        ("BI-24-02", "lecture"),
        ("BI-24-02", "practical"),
    }


def test_schedule_generation_success_flow_with_export(client, admin_auth_headers):
    _seed_schedule_data(client, admin_auth_headers)

    generate_response = client.post(
        "/api/schedules/generate",
        headers=admin_auth_headers,
        json={"semester": 1, "year": 2026, "algorithm": "optimizer"},
    )

    assert generate_response.status_code == 202
    job_id = generate_response.json()["jobId"]

    final_job = _wait_for_job_completion(client, admin_auth_headers, job_id)
    assert final_job["status"] == "completed"
    assert final_job["result"]["scheduleCount"] == 2

    schedules_response = client.get("/api/schedules", headers=admin_auth_headers)
    assert schedules_response.status_code == 200
    schedules = schedules_response.json()
    assert len(schedules) == 2
    assert {item["course_name"] for item in schedules} == {"Algorithms"}
    assert {item["group_name"] for item in schedules} == {"SE-24-01"}
    assert {item["teacher_name"] for item in schedules} == {"Aruzhan Saparova"}
    assert {item["room_number"] for item in schedules} == {"401"}
    assert {item["semester"] for item in schedules} == {1}
    assert {item["year"] for item in schedules} == {2026}

    export_response = client.get("/api/export/schedule", headers=admin_auth_headers)
    assert export_response.status_code == 200

    workbook = load_workbook(filename=BytesIO(export_response.content), data_only=True)
    sheet = workbook["Schedule"]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == (
        "course_name",
        "group_name",
        "subgroup",
        "teacher_name",
        "room_number",
        "day",
        "start_hour",
        "semester",
        "year",
        "algorithm",
        "room_programme",
        "room_programme_mismatch",
    )
    assert len(rows) == 3
    assert all(row[0] == "Algorithms" for row in rows[1:])
    assert all(row[1] == "SE-24-01" for row in rows[1:])
    assert all(row[4] == "401" for row in rows[1:])


def test_manual_schedule_rejects_teacher_conflict(client, admin_auth_headers):
    _seed_schedule_data(client, admin_auth_headers)

    sections_response = client.get("/api/sections", headers=admin_auth_headers)
    rooms_response = client.get("/api/rooms", headers=admin_auth_headers)
    assert sections_response.status_code == 200
    assert rooms_response.status_code == 200
    section = sections_response.json()[0]
    room = rooms_response.json()[0]

    payload = {
        "section_id": section["id"],
        "course_id": section["course_id"],
        "course_name": section["course_name"],
        "teacher_id": section["teacher_id"],
        "teacher_name": section["teacher_name"],
        "room_id": room["id"],
        "room_number": room["number"],
        "group_id": section["group_id"],
        "group_name": section["group_name"],
        "subgroup": "",
        "day": "2026-04-20",
        "start_hour": 9,
        "semester": 1,
        "year": 2026,
        "algorithm": "manual",
    }

    first_response = client.post("/api/schedules", headers=admin_auth_headers, json=payload)
    assert first_response.status_code == 201

    second_response = client.post("/api/schedules", headers=admin_auth_headers, json=payload)
    assert second_response.status_code == 400
    assert second_response.json()["errorCode"] == "bad_request"


def test_manual_schedule_keeps_room_active_but_blocks_the_occupied_slot(client, admin_auth_headers):
    _seed_schedule_data(client, admin_auth_headers)

    sections_response = client.get("/api/sections", headers=admin_auth_headers)
    rooms_response = client.get("/api/rooms", headers=admin_auth_headers)
    assert sections_response.status_code == 200
    assert rooms_response.status_code == 200
    section = sections_response.json()[0]
    room = rooms_response.json()[0]

    payload = {
        "section_id": section["id"],
        "course_id": section["course_id"],
        "course_name": section["course_name"],
        "teacher_id": section["teacher_id"],
        "teacher_name": section["teacher_name"],
        "room_id": room["id"],
        "room_number": room["number"],
        "group_id": section["group_id"],
        "group_name": section["group_name"],
        "subgroup": "",
        "day": "2026-04-20",
        "start_hour": 9,
        "semester": 1,
        "year": 2026,
        "algorithm": "manual",
    }

    create_response = client.post("/api/schedules", headers=admin_auth_headers, json=payload)
    assert create_response.status_code == 201

    rooms_after_create = client.get("/api/rooms", headers=admin_auth_headers)
    assert rooms_after_create.status_code == 200
    assert rooms_after_create.json()[0]["available"] == 1

    conflicting_create_response = client.post(
        "/api/schedules",
        headers=admin_auth_headers,
        json=payload,
    )
    assert conflicting_create_response.status_code == 400
    assert conflicting_create_response.json()["error"] == "Аудитория уже занята в это время"

    reset_response = client.post(
        "/api/schedules/reset",
        headers=admin_auth_headers,
        json={"semester": 1, "year": 2026},
    )
    assert reset_response.status_code == 200

    rooms_after_reset = client.get("/api/rooms", headers=admin_auth_headers)
    assert rooms_after_reset.status_code == 200
    assert rooms_after_reset.json()[0]["available"] == 1


def test_schedule_export_supports_semester_and_year_scope(client, admin_auth_headers):
    _seed_schedule_data(client, admin_auth_headers)

    sections_response = client.get("/api/sections", headers=admin_auth_headers)
    rooms_response = client.get("/api/rooms", headers=admin_auth_headers)
    assert sections_response.status_code == 200
    assert rooms_response.status_code == 200
    section = sections_response.json()[0]
    room = rooms_response.json()[0]

    first_payload = {
        "section_id": section["id"],
        "course_id": section["course_id"],
        "course_name": section["course_name"],
        "teacher_id": section["teacher_id"],
        "teacher_name": section["teacher_name"],
        "room_id": room["id"],
        "room_number": room["number"],
        "group_id": section["group_id"],
        "group_name": section["group_name"],
        "subgroup": "",
        "day": "2026-04-20",
        "start_hour": 9,
        "semester": 1,
        "year": 2026,
        "algorithm": "manual",
    }
    second_payload = {
        **first_payload,
        "day": "2026-04-21",
        "start_hour": 10,
        "semester": 2,
        "year": 2027,
    }

    assert client.post("/api/schedules", headers=admin_auth_headers, json=first_payload).status_code == 201
    assert client.post("/api/schedules", headers=admin_auth_headers, json=second_payload).status_code == 201

    export_response = client.get(
        "/api/export/schedule?semester=1&year=2026",
        headers=admin_auth_headers,
    )
    assert export_response.status_code == 200

    workbook = load_workbook(filename=BytesIO(export_response.content), data_only=True)
    sheet = workbook["Schedule"]
    rows = list(sheet.iter_rows(values_only=True))

    assert len(rows) == 2
    assert rows[1][0] == "Algorithms"
    assert rows[1][7] == 1
    assert rows[1][8] == 2026


def test_optimizer_allows_parallel_different_subgroups():
    from backend.app.optimizer import optimize_schedule

    result = optimize_schedule(
        {
            "timeSlots": [{"id": "monday_8", "day": "Monday", "hour": 8}],
            "teachers": [
                {"id": 1, "name": "Teacher A"},
                {"id": 2, "name": "Teacher B"},
            ],
            "rooms": [
                {"id": 1, "number": "101", "capacity": 20, "type": "practical", "pcCount": 20},
                {"id": 2, "number": "102", "capacity": 20, "type": "practical", "pcCount": 20},
            ],
            "planItems": [
                {
                    "id": "lab_a",
                    "courseId": 1,
                    "courseName": "Programming",
                    "teacherId": 1,
                    "groupIds": ["G1"],
                    "subgroupIds": ["G1-A"],
                    "lessonsPerWeek": 1,
                    "studentCount": 15,
                    "lessonType": "lab",
                    "roomTypeRequired": "practical",
                    "pcRequired": True,
                },
                {
                    "id": "lab_b",
                    "courseId": 1,
                    "courseName": "Programming",
                    "teacherId": 2,
                    "groupIds": ["G1"],
                    "subgroupIds": ["G1-B"],
                    "lessonsPerWeek": 1,
                    "studentCount": 15,
                    "lessonType": "lab",
                    "roomTypeRequired": "practical",
                    "pcRequired": True,
                },
            ],
        }
    )

    assert result["status"] in {"OPTIMAL", "FEASIBLE"}
    assert len(result["schedule"]) == 2
    assert {item["hour"] for item in result["schedule"]} == {8}


def test_optimizer_allows_lab_in_practical_room_with_at_least_ten_computers():
    from backend.app.optimizer import optimize_schedule

    result = optimize_schedule(
        {
            "timeSlots": [{"id": "monday_8", "day": "Monday", "hour": 8}],
            "teachers": [
                {"id": 1, "name": "Teacher A"},
            ],
            "rooms": [
                {"id": 1, "number": "201", "capacity": 30, "type": "practical", "pcCount": 12},
            ],
            "planItems": [
                {
                    "id": "lab_main",
                    "courseId": 1,
                    "courseName": "Databases",
                    "teacherId": 1,
                    "groupIds": ["G1"],
                    "lessonsPerWeek": 1,
                    "studentCount": 20,
                    "lessonType": "lab",
                    "roomTypeRequired": "practical",
                    "pcRequired": True,
                },
            ],
        }
    )

    assert result["status"] in {"OPTIMAL", "FEASIBLE"}
    assert len(result["schedule"]) == 1
    assert result["schedule"][0]["roomId"] == 1


def test_optimizer_rejects_computer_lesson_when_practical_room_has_less_than_ten_computers():
    from backend.app.optimizer import optimize_schedule
    from backend.app.errors import ApiError

    with pytest.raises(ApiError) as exc_info:
        optimize_schedule(
            {
                "timeSlots": [{"id": "monday_8", "day": "Monday", "hour": 8}],
                "teachers": [
                    {"id": 1, "name": "Teacher A"},
                ],
                "rooms": [
                    {"id": 1, "number": "202", "capacity": 30, "type": "practical", "pcCount": 9},
                ],
                "planItems": [
                    {
                        "id": "lab_main",
                        "courseId": 1,
                        "courseName": "Databases",
                        "teacherId": 1,
                        "groupIds": ["G1"],
                        "lessonsPerWeek": 1,
                        "studentCount": 20,
                        "lessonType": "lab",
                        "roomTypeRequired": "practical",
                        "pcRequired": True,
                    },
                ],
            }
        )

    assert exc_info.value.code == "optimizer_input_infeasible"
